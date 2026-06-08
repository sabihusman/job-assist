"""Read-endpoint tests for PR #30a.

Sync tests cover request-validation; DB-gated tests cover the shape +
filter + pagination + ordering + no-N+1 behaviour of:
  GET /postings           — list
  GET /postings/{id}      — detail
  GET /companies          — list with counts
  GET /outcomes           — chronological event list
"""

from __future__ import annotations

import os
import uuid
from datetime import UTC, datetime, timedelta
from typing import Any

import pytest
from httpx import ASGITransport, AsyncClient

from job_assist.db.models import (
    Division,
    JobPosting,
    OutcomeEvent,
    PostingSource,
    TargetCompany,
)

_NEEDS_DB = pytest.mark.skipif(
    not os.getenv("TEST_DATABASE_URL"),
    reason="TEST_DATABASE_URL not set",
)


# ── ASGI client + query-count helper ─────────────────────────────────────────


async def _client(db_session: Any) -> AsyncClient:
    from job_assist.db.session import get_db
    from job_assist.main import app

    async def _override() -> Any:
        yield db_session

    app.dependency_overrides[get_db] = _override
    return AsyncClient(transport=ASGITransport(app=app), base_url="http://test")


async def _drop_override() -> None:
    from job_assist.db.session import get_db
    from job_assist.main import app

    app.dependency_overrides.pop(get_db, None)


class _ExecuteCounter:
    """Wraps ``session.execute`` to count how many SQL statements the
    endpoint emits. The tests assert ≤ N for the no-N+1 contract.

    Implements both sync and async context-manager protocols so it can be
    used with ``with counter:`` OR ``async with counter:`` — convenient
    when nesting alongside other ``async with`` blocks (like httpx's
    ``AsyncClient``).
    """

    def __init__(self, session: Any) -> None:
        self._session = session
        self._original = session.execute
        self.count = 0

    async def _wrapped(self, *args: Any, **kwargs: Any) -> Any:
        self.count += 1
        return await self._original(*args, **kwargs)

    def __enter__(self) -> _ExecuteCounter:
        self._session.execute = self._wrapped  # type: ignore[method-assign]
        return self

    def __exit__(self, *_exc: Any) -> None:
        self._session.execute = self._original  # type: ignore[method-assign]

    async def __aenter__(self) -> _ExecuteCounter:
        return self.__enter__()

    async def __aexit__(self, *_exc: Any) -> None:
        self.__exit__(*_exc)


# ── Factories ────────────────────────────────────────────────────────────────


def _company(
    *,
    name: str,
    tier: int = 1,
    ats: str = "greenhouse",
    domain: str | None = None,
    description: str | None = None,
) -> TargetCompany:
    return TargetCompany(
        name=name,
        tier=tier,
        ats=ats,
        ats_handle=f"handle-{uuid.uuid4().hex[:6]}",
        domain=domain,
        description=description,
    )


def _posting(
    *,
    target_company_id: uuid.UUID | None,
    title: str = "Senior Product Manager",
    role_family: str = "product_management",
    remote_type: str = "remote",
    department: str | None = None,
    team: str | None = None,
    salary_max: int | None = None,
    closed: bool = False,
    hard_rule_failed: str | None = None,
    first_seen_at: datetime | None = None,
    fit_score: int | None = None,
    similarity_score: int | None = None,
) -> JobPosting:
    now = first_seen_at or datetime.now(tz=UTC)
    suffix = uuid.uuid4().hex[:10]
    return JobPosting(
        canonical_company_name="TestCo",
        target_company_id=target_company_id,
        normalized_title=title.lower(),
        raw_title=title,
        remote_type=remote_type,
        role_family=role_family,
        seniority_level="senior_pm",
        jd_text="JD body.",
        jd_text_hash="0" * 64,
        content_hash=f"hash-{suffix}",
        first_seen_at=now,
        last_seen_at=now,
        salary_max=salary_max,
        department=department,
        team=team,
        closed_at=now if closed else None,
        hard_rule_failed=hard_rule_failed,
        fit_score=fit_score,
        similarity_score=similarity_score,
    )


def _posting_source(
    *,
    job_posting_id: uuid.UUID,
    ats: str = "greenhouse",
    url: str | None = None,
    fetched_at: datetime | None = None,
) -> PostingSource:
    return PostingSource(
        job_posting_id=job_posting_id,
        ats=ats,
        source_job_id=uuid.uuid4().hex,
        source_url=url or f"https://jobs.example.com/{uuid.uuid4().hex[:8]}",
        apply_url=None,
        raw_payload={},
        parser_version="test-v1",
        fetch_status="ok",
        fetched_at=fetched_at or datetime.now(tz=UTC),
    )


# ── /postings — sync request validation ──────────────────────────────────────


@_NEEDS_DB
async def test_postings_limit_too_large(db_session: Any) -> None:
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?limit=200")
    finally:
        await _drop_override()
    assert resp.status_code == 422


@_NEEDS_DB
async def test_postings_invalid_ats_rejected(db_session: Any) -> None:
    ac = await _client(db_session)
    try:
        async with ac:
            # ``taleo`` is a real ATS but not one this app ingests, so it's not
            # in _ALLOWED_ATS_VALUES. (``workday``/``icims`` ARE valid now —
            # fix/datacenter-egress-headers added them to the filter allowlist.)
            resp = await ac.get("/postings?ats=taleo")
    finally:
        await _drop_override()
    assert resp.status_code == 422
    assert "taleo" in resp.text


@_NEEDS_DB
async def test_postings_invalid_remote_type_rejected(db_session: Any) -> None:
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?remote_type=hybrid_remote")
    finally:
        await _drop_override()
    assert resp.status_code == 422


# ── /postings — DB-backed ────────────────────────────────────────────────────


@_NEEDS_DB
async def test_postings_list_empty(db_session: Any) -> None:
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings")
    finally:
        await _drop_override()
    assert resp.status_code == 200
    body = resp.json()
    assert body == {"total": 0, "offset": 0, "limit": 20, "items": []}


@_NEEDS_DB
async def test_postings_default_pagination_and_total(db_session: Any) -> None:
    company = _company(name="PaginateCo")
    db_session.add(company)
    await db_session.flush()

    # 30 postings, each ~1 minute apart so the sort is deterministic.
    base = datetime(2026, 5, 1, tzinfo=UTC)
    for i in range(30):
        jp = _posting(target_company_id=company.id, first_seen_at=base + timedelta(minutes=i))
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            # Disable per-company cap — all 30 rows belong to PaginateCo
            # and the cap (default 3, PR #58) would collapse them.
            # This test's invariant is pagination math.
            resp = await ac.get("/postings?per_company_cap=0")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 30
    assert body["offset"] == 0
    assert body["limit"] == 20
    assert len(body["items"]) == 20


@_NEEDS_DB
async def test_postings_offset_pagination(db_session: Any) -> None:
    company = _company(name="OffsetCo")
    db_session.add(company)
    await db_session.flush()
    base = datetime(2026, 5, 1, tzinfo=UTC)
    for i in range(15):
        jp = _posting(target_company_id=company.id, first_seen_at=base + timedelta(minutes=i))
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            # Disable per-company cap (PR #58 default=3) — fixture is
            # 15 postings from one company and the cap would mask
            # the pagination math under test.
            resp = await ac.get("/postings?limit=5&offset=10&per_company_cap=0")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 15
    assert body["limit"] == 5
    assert body["offset"] == 10
    assert len(body["items"]) == 5


@_NEEDS_DB
async def test_postings_filter_by_tier(db_session: Any) -> None:
    c1 = _company(name="T1Co", tier=1)
    c2 = _company(name="T3Co", tier=3)
    db_session.add_all([c1, c2])
    await db_session.flush()
    for c in (c1, c2):
        jp = _posting(target_company_id=c.id)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?tier=1")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 1
    assert body["items"][0]["company"]["name"] == "T1Co"


@_NEEDS_DB
async def test_postings_filter_by_tier_multi(db_session: Any) -> None:
    c1 = _company(name="MultiT1", tier=1)
    c2 = _company(name="MultiT2", tier=2)
    c3 = _company(name="MultiT3", tier=3)
    db_session.add_all([c1, c2, c3])
    await db_session.flush()
    for c in (c1, c2, c3):
        jp = _posting(target_company_id=c.id)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?tier=1&tier=2")
    finally:
        await _drop_override()

    names = {item["company"]["name"] for item in resp.json()["items"]}
    assert names == {"MultiT1", "MultiT2"}


@_NEEDS_DB
async def test_postings_filter_by_ats(db_session: Any) -> None:
    c = _company(name="AtsFilterCo")
    db_session.add(c)
    await db_session.flush()
    jp_gh = _posting(target_company_id=c.id)
    jp_lev = _posting(target_company_id=c.id)
    db_session.add_all([jp_gh, jp_lev])
    await db_session.flush()
    db_session.add_all(
        [
            _posting_source(job_posting_id=jp_gh.id, ats="greenhouse"),
            _posting_source(job_posting_id=jp_lev.id, ats="lever"),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?ats=greenhouse")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    assert {item["source"]["ats"] for item in items} == {"greenhouse"}


@_NEEDS_DB
async def test_postings_filter_by_remote_type(db_session: Any) -> None:
    c = _company(name="RemoteFilterCo")
    db_session.add(c)
    await db_session.flush()
    db_session.add_all(
        [
            _posting(target_company_id=c.id, remote_type="remote"),
            _posting(target_company_id=c.id, remote_type="onsite"),
        ]
    )
    await db_session.flush()
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?remote_type=remote")
    finally:
        await _drop_override()

    assert {item["remote_type"] for item in resp.json()["items"]} == {"remote"}


@_NEEDS_DB
async def test_postings_filter_by_role_family_case_insensitive(db_session: Any) -> None:
    c = _company(name="RoleFamCo")
    db_session.add(c)
    await db_session.flush()
    db_session.add_all(
        [
            _posting(target_company_id=c.id, role_family="product_management"),
            _posting(target_company_id=c.id, role_family="product_marketing"),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            # Uppercase input must still match the lowercase enum value.
            resp = await ac.get("/postings?role_family=PRODUCT_MANAGEMENT")
    finally:
        await _drop_override()

    assert all(item["role"]["family"] == "product_management" for item in resp.json()["items"])


@_NEEDS_DB
async def test_postings_filter_by_target_company_id(db_session: Any) -> None:
    c1 = _company(name="TargetA")
    c2 = _company(name="TargetB")
    db_session.add_all([c1, c2])
    await db_session.flush()
    db_session.add_all(
        [
            _posting(target_company_id=c1.id),
            _posting(target_company_id=c2.id),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings?target_company_id={c1.id}")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    assert len(items) == 1
    assert items[0]["company"]["id"] == str(c1.id)


@_NEEDS_DB
async def test_postings_sorted_first_seen_desc(db_session: Any) -> None:
    c = _company(name="SortedCo")
    db_session.add(c)
    await db_session.flush()
    base = datetime(2026, 5, 1, tzinfo=UTC)
    db_session.add_all(
        [
            _posting(target_company_id=c.id, first_seen_at=base),
            _posting(target_company_id=c.id, first_seen_at=base + timedelta(days=1)),
            _posting(target_company_id=c.id, first_seen_at=base + timedelta(days=2)),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings")
    finally:
        await _drop_override()

    ts = [item["first_seen_at"] for item in resp.json()["items"]]
    assert ts == sorted(ts, reverse=True)


@_NEEDS_DB
async def test_postings_response_shape(db_session: Any) -> None:
    """Golden shape on one fully-populated row."""
    c = _company(
        name="ShapeCo",
        tier=1,
        domain="shape.example",
        description="ShapeCo builds shapes.",
    )
    db_session.add(c)
    await db_session.flush()
    jp = _posting(
        target_company_id=c.id,
        title="Senior PM",
        department="Product",
        team="Risk",
        salary_max=180_000,
    )
    db_session.add(jp)
    await db_session.flush()
    db_session.add(
        _posting_source(
            job_posting_id=jp.id,
            ats="greenhouse",
            url="https://jobs.shape.example/123",
        )
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings")
    finally:
        await _drop_override()

    item = resp.json()["items"][0]
    assert set(item) == {
        "id",
        "company",
        "role",
        "location_raw",
        "locations_normalized",
        "remote_type",
        "salary",
        "source",
        "first_seen_at",
        "score",
        # Slice 2b: calibrated 0-100 semantic similarity (null until recalibrated).
        "similarity_score",
        # Added in PR #31 — always present, nested fields null for untouched postings.
        "state",
    }
    assert item["company"]["name"] == "ShapeCo"
    assert item["company"]["domain"] == "shape.example"
    assert item["company"]["tier"] == 1
    assert item["role"]["department"] == "Product"
    assert item["role"]["team"] == "Risk"
    assert item["salary"]["max"] == 180_000
    assert item["source"]["ats"] == "greenhouse"
    assert item["source"]["url"] == "https://jobs.shape.example/123"
    assert item["score"] is None


@_NEEDS_DB
async def test_postings_no_n_plus_one(db_session: Any) -> None:
    """20 postings across 5 companies → 2 SQL statements total.

    PR #58: the default ``per_company_cap=3`` would clip this fixture
    from 20 to 15 (each company has 4 postings; cap takes 3). The
    no-N+1 test predates the cap and wants to count ALL 20 rows to
    prove the lateral joins don't multiply queries — we explicitly
    disable the cap here so the test's invariant survives the new
    default. The cap itself is covered by tests/test_per_company_cap.py.
    """
    companies = [_company(name=f"NPlusOne{i}", tier=i % 4 + 1) for i in range(5)]
    db_session.add_all(companies)
    await db_session.flush()
    for c in companies:
        for _ in range(4):
            jp = _posting(target_company_id=c.id)
            db_session.add(jp)
            await db_session.flush()
            db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    counter = _ExecuteCounter(db_session)
    ac = await _client(db_session)
    try:
        async with ac, counter:
            resp = await ac.get("/postings?per_company_cap=0")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert resp.json()["total"] == 20
    assert counter.count <= 2, f"expected ≤2 queries, got {counter.count}"


@_NEEDS_DB
async def test_postings_excludes_closed_by_default(db_session: Any) -> None:
    """Stale-posting filter (Bestiary 5.18): closed postings are hidden by
    default; ``include_closed=true`` brings them back; ``total`` (COUNT)
    respects the same filter as the page (SELECT)."""
    c = _company(name="ClosedFilterCo")
    db_session.add(c)
    await db_session.flush()
    # 2 open + 3 closed, same company.
    for closed in (False, False, True, True, True):
        jp = _posting(target_company_id=c.id, closed=closed)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            # Default: only the 2 open postings, and total counts only those.
            default_resp = await ac.get("/postings?per_company_cap=0")
            # include_closed=true: all 5 surface, total counts all.
            all_resp = await ac.get("/postings?per_company_cap=0&include_closed=true")
    finally:
        await _drop_override()

    default_body = default_resp.json()
    assert default_body["total"] == 2, "COUNT must exclude closed by default"
    assert len(default_body["items"]) == 2

    all_body = all_resp.json()
    assert all_body["total"] == 5, "include_closed=true must count closed too"
    assert len(all_body["items"]) == 5


@_NEEDS_DB
async def test_postings_excludes_hard_rule_failed_by_default(db_session: Any) -> None:
    """PR C: postings that failed a hard rule are hidden by default;
    ``include_filtered=true`` brings them back; COUNT respects the filter.
    NULL ``hard_rule_failed`` (passed or not-yet-evaluated) always surfaces."""
    c = _company(name="HardRuleFilterCo")
    db_session.add(c)
    await db_session.flush()
    # 2 passing (NULL) + 3 filtered (failed a rule), same company.
    for failed in (None, None, "salary_floor", "geo_whitelist", "salary_ceiling"):
        jp = _posting(target_company_id=c.id, hard_rule_failed=failed)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            default_resp = await ac.get("/postings?per_company_cap=0")
            all_resp = await ac.get("/postings?per_company_cap=0&include_filtered=true")
    finally:
        await _drop_override()

    default_body = default_resp.json()
    assert default_body["total"] == 2, "COUNT must exclude hard-rule-failed by default"
    assert len(default_body["items"]) == 2

    all_body = all_resp.json()
    assert all_body["total"] == 5, "include_filtered=true must count filtered too"
    assert len(all_body["items"]) == 5


@_NEEDS_DB
async def test_postings_closed_and_hard_rule_filters_compose(db_session: Any) -> None:
    """The closed_at and hard_rule_failed filters stack via AND — a posting
    must be BOTH open AND pass-hard-rules to show by default."""
    c = _company(name="ComposeCo")
    db_session.add(c)
    await db_session.flush()
    # Only the first is open AND passing → the only default-visible row.
    specs = [
        (False, None),  # open + passing  ✓ visible
        (True, None),  # closed + passing  ✗
        (False, "salary_floor"),  # open + filtered  ✗
        (True, "geo_whitelist"),  # closed + filtered ✗
    ]
    for closed, failed in specs:
        jp = _posting(target_company_id=c.id, closed=closed, hard_rule_failed=failed)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=0")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 1, "only the open+passing posting shows by default"
    assert len(body["items"]) == 1


@_NEEDS_DB
async def test_reeval_hard_rules_endpoint(db_session: Any) -> None:
    """POST /admin/postings/reeval-hard-rules re-evaluates open postings and
    rewrites hard_rule_failed. Seeds the operator_profile (singleton) so the
    mapper has a row to read."""
    from sqlalchemy import select

    from job_assist.db.models import OperatorProfile

    # Upsert the singleton — the test DB is already seeded with id=1, so a
    # plain insert would violate the PK. Overwrite the rule fields the test
    # depends on so it's deterministic regardless of seeded defaults.
    profile = (
        await db_session.execute(select(OperatorProfile).where(OperatorProfile.id == 1))
    ).scalar_one_or_none()
    fields = {
        "geo_whitelist": ["Remote"],
        "salary_floor_usd": 150_000,
        "salary_ceiling_usd": None,
        "applicant_cap": 500,
        "seniority_levels_included": None,
        "staffing_firm_blocklist": [],
    }
    if profile is None:
        db_session.add(OperatorProfile(id=1, looking_for_text="PM", role_keywords=[], **fields))
    else:
        for key, value in fields.items():
            setattr(profile, key, value)
    c = _company(name="ReevalCo")
    db_session.add(c)
    await db_session.flush()
    # below-floor (disclosed) → should fail salary_floor;
    # null-salary → passes; closed → not evaluated (stays as-is).
    below = _posting(target_company_id=c.id, salary_max=90_000)
    below.salary_currency = "USD"
    below.salary_period = "annual"
    nullsal = _posting(target_company_id=c.id, salary_max=None)
    closed = _posting(target_company_id=c.id, salary_max=80_000, closed=True)
    closed.salary_currency = "USD"
    closed.salary_period = "annual"
    db_session.add_all([below, nullsal, closed])
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.post("/admin/postings/reeval-hard-rules")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    body = resp.json()
    assert body["evaluated"] == 2, "only the 2 OPEN postings are evaluated"
    assert body["passed"] == 1
    assert body["failed"] == 1
    assert body["by_rule"].get("salary_floor") == 1

    await db_session.refresh(below)
    await db_session.refresh(nullsal)
    assert below.hard_rule_failed == "salary_floor"
    assert nullsal.hard_rule_failed is None


# ── reparse-salary backfill (PR feat/reparse-salary-backfill) ────────────────


@_NEEDS_DB
async def test_reparse_salary_corrects_inverted_overwrites_garbage_no_churn(
    db_session: Any,
) -> None:
    """POST /admin/postings/reparse-salary on Greenhouse-sourced open postings:
    * inverted range (old min > max from a US+CAD JD)  → corrected
    * $142M-style garbage (no other range in body)     → set to NULL
    * already-correct row                              → no churn
    """
    c = _company(name="ReparseCo")
    db_session.add(c)
    await db_session.flush()

    # Real-world failure-mode source text (audit findings).
    inverted_jd = (
        "compensation (any location): $189,000-236,200 USD for US employees "
        "outside SF, and $178,600-223,200 CAD for Canada."
    )
    # Garbled $142M with no other range in the body — new parser rejects to None.
    garbage_jd = "Comp for this role is $142,400,000 base. No other range listed."
    # A correctly-stored row whose jd_text parses to the same values it has.
    clean_jd = "Base pay: $180,000 - $275,000 USD plus equity."

    inverted = _posting(target_company_id=c.id, salary_max=178_600)
    inverted.jd_text = inverted_jd
    inverted.salary_min = 189_000  # old wrong value the parser will fix
    inverted.salary_currency = "USD"
    inverted.salary_period = "annual"  # type: ignore[assignment]

    garbage = _posting(target_company_id=c.id, salary_max=178_000)
    garbage.jd_text = garbage_jd
    garbage.salary_min = 142_400_000  # old garbage the parser rejects
    garbage.salary_currency = "USD"
    garbage.salary_period = "annual"  # type: ignore[assignment]

    clean = _posting(target_company_id=c.id, salary_max=275_000)
    clean.jd_text = clean_jd
    clean.salary_min = 180_000  # already correct
    clean.salary_currency = "USD"
    clean.salary_period = "annual"  # type: ignore[assignment]

    db_session.add_all([inverted, garbage, clean])
    await db_session.flush()
    db_session.add_all(
        [
            _posting_source(job_posting_id=inverted.id, ats="greenhouse"),
            _posting_source(job_posting_id=garbage.id, ats="greenhouse"),
            _posting_source(job_posting_id=clean.id, ats="greenhouse"),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.post("/admin/postings/reparse-salary")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    body = resp.json()
    assert body["evaluated"] == 3
    assert body["changed"] == 2  # inverted + garbage; clean stays
    assert body["inversions_fixed"] == 1
    assert body["rejected_to_null"] == 1

    await db_session.refresh(inverted)
    await db_session.refresh(garbage)
    await db_session.refresh(clean)

    # Inverted → US range (both ends), ordered.
    assert inverted.salary_min == 189_000
    assert inverted.salary_max == 236_200
    assert inverted.salary_currency == "USD"
    assert inverted.salary_min <= inverted.salary_max

    # Garbage → nulled (no plausible range).
    assert garbage.salary_min is None
    assert garbage.salary_max is None
    assert garbage.salary_currency is None

    # Clean → unchanged.
    assert clean.salary_min == 180_000
    assert clean.salary_max == 275_000


@_NEEDS_DB
async def test_reparse_salary_skips_non_greenhouse_sources(
    db_session: Any,
) -> None:
    """Only Greenhouse-sourced postings are reparsed — Ashby feeds a clean
    compensationTierSummary to the parser, not the JD body, so it doesn't
    have this failure mode. An Ashby posting with stored garbage is left
    alone by this endpoint (its correctness lives in the adapter)."""
    c = _company(name="AshbySkipCo")
    db_session.add(c)
    await db_session.flush()

    ashby = _posting(target_company_id=c.id, salary_max=178_000)
    ashby.jd_text = "Whatever JD body, $142,400,000 stray."
    ashby.salary_min = 142_400_000
    ashby.salary_currency = "USD"
    ashby.salary_period = "annual"  # type: ignore[assignment]
    db_session.add(ashby)
    await db_session.flush()
    db_session.add(_posting_source(job_posting_id=ashby.id, ats="ashby"))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.post("/admin/postings/reparse-salary")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert resp.json()["evaluated"] == 0  # ashby row not in scope

    await db_session.refresh(ashby)
    assert ashby.salary_min == 142_400_000  # untouched


# ── /postings/{id} ───────────────────────────────────────────────────────────


@_NEEDS_DB
async def test_posting_detail_full_shape(db_session: Any) -> None:
    c = _company(name="DetailCo", description="DetailCo description")
    db_session.add(c)
    await db_session.flush()
    jp = _posting(target_company_id=c.id, department="Eng", team="Platform")
    db_session.add(jp)
    await db_session.flush()
    db_session.add(_posting_source(job_posting_id=jp.id))
    db_session.add(
        Division(
            target_company_id=c.id,
            department="Eng",
            team="Platform",
            description="The Eng/Platform division ships infra.",
        )
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings/{jp.id}")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    body = resp.json()
    assert body["id"] == str(jp.id)
    assert body["description_markdown"] == "JD body."
    assert body["division"] is not None
    assert body["division"]["department"] == "Eng"
    assert body["division"]["team"] == "Platform"
    assert "infra" in body["division"]["description"]


@_NEEDS_DB
async def test_posting_detail_404_unknown_id(db_session: Any) -> None:
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings/{uuid.uuid4()}")
    finally:
        await _drop_override()
    assert resp.status_code == 404


@_NEEDS_DB
async def test_posting_detail_division_null_when_no_match(db_session: Any) -> None:
    c = _company(name="NoDivCo")
    db_session.add(c)
    await db_session.flush()
    jp = _posting(target_company_id=c.id, department="Eng", team=None)
    db_session.add(jp)
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings/{jp.id}")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert resp.json()["division"] is None


@_NEEDS_DB
async def test_posting_detail_division_matches_via_nulls(db_session: Any) -> None:
    """Posting (Eng, NULL) + division (Eng, NULL) match via IS NOT DISTINCT FROM."""
    c = _company(name="NullsMatch")
    db_session.add(c)
    await db_session.flush()
    jp = _posting(target_company_id=c.id, department="Eng", team=None)
    db_session.add(jp)
    db_session.add(
        Division(
            target_company_id=c.id,
            department="Eng",
            team=None,
            description="The Eng division",
        )
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings/{jp.id}")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["division"] is not None
    assert body["division"]["department"] == "Eng"
    assert body["division"]["team"] is None


# ── /companies ───────────────────────────────────────────────────────────────


@_NEEDS_DB
async def test_companies_list_returns_counts(db_session: Any) -> None:
    c1 = _company(name="CountsA", tier=1)
    c2 = _company(name="CountsB", tier=2)
    c3 = _company(name="CountsC", tier=3)
    db_session.add_all([c1, c2, c3])
    await db_session.flush()
    for _ in range(5):
        db_session.add(_posting(target_company_id=c1.id))
    for _ in range(3):
        db_session.add(_posting(target_company_id=c2.id))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    by_name = {c["name"]: c for c in resp.json()["items"]}
    assert by_name["CountsA"]["total_postings"] == 5
    assert by_name["CountsB"]["total_postings"] == 3
    assert by_name["CountsC"]["total_postings"] == 0


@_NEEDS_DB
async def test_companies_active_vs_total(db_session: Any) -> None:
    c = _company(name="ActiveVsTotal")
    db_session.add(c)
    await db_session.flush()
    for closed in (False, False, False, True, True):
        db_session.add(_posting(target_company_id=c.id, closed=closed))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    row = next(c for c in resp.json()["items"] if c["name"] == "ActiveVsTotal")
    assert row["total_postings"] == 5
    assert row["active_postings"] == 3


@_NEEDS_DB
async def test_companies_ats_set(db_session: Any) -> None:
    c = _company(name="AtsSetCo")
    db_session.add(c)
    await db_session.flush()
    jp1 = _posting(target_company_id=c.id)
    jp2 = _posting(target_company_id=c.id)
    db_session.add_all([jp1, jp2])
    await db_session.flush()
    db_session.add_all(
        [
            _posting_source(job_posting_id=jp1.id, ats="greenhouse"),
            _posting_source(job_posting_id=jp2.id, ats="lever"),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    row = next(c for c in resp.json()["items"] if c["name"] == "AtsSetCo")
    assert set(row["ats_set"]) == {"greenhouse", "lever"}


@_NEEDS_DB
async def test_companies_filter_by_tier(db_session: Any) -> None:
    c1 = _company(name="CompTierA", tier=1)
    c2 = _company(name="CompTierB", tier=3)
    db_session.add_all([c1, c2])
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies?tier=3")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    assert {c["name"] for c in items} == {"CompTierB"}


@_NEEDS_DB
async def test_companies_sort_tier_then_name(db_session: Any) -> None:
    db_session.add_all(
        [
            _company(name="Zeta T2", tier=2),
            _company(name="Alpha T3", tier=3),
            _company(name="Mid T1", tier=1),
            _company(name="Bravo T1", tier=1),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    names = [c["name"] for c in resp.json()["items"]]
    # Filter to the four we just inserted so existing fixtures don't pollute.
    ours = [n for n in names if n in {"Bravo T1", "Mid T1", "Zeta T2", "Alpha T3"}]
    assert ours == ["Bravo T1", "Mid T1", "Zeta T2", "Alpha T3"]


@_NEEDS_DB
async def test_companies_no_n_plus_one(db_session: Any) -> None:
    """5 companies x 10 postings each -> 2 SQL statements total."""
    companies = [_company(name=f"NPlus1Co{i}", tier=i % 4 + 1) for i in range(5)]
    db_session.add_all(companies)
    await db_session.flush()
    for c in companies:
        for _ in range(10):
            jp = _posting(target_company_id=c.id)
            db_session.add(jp)
            await db_session.flush()
            db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    counter = _ExecuteCounter(db_session)
    ac = await _client(db_session)
    try:
        async with ac, counter:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert counter.count <= 2, f"expected ≤2 queries, got {counter.count}"


# ── PR #71: paused-state surfacing ─────────────────────────────────────────


@_NEEDS_DB
async def test_companies_surfaces_ats_handle_and_notes_pr71(db_session: Any) -> None:
    """Companies response must include ``ats``, ``ats_handle``, ``notes``.

    The frontend Companies table (PR #71) uses these to render a Paused
    badge when an operator soft-pauses a target (PR #65 Atlassian case
    — adapter known, handle cleared, notes field explains why).
    """
    live = TargetCompany(
        name="LiveCo_PR71",
        tier=1,
        ats="greenhouse",
        ats_handle="livehandle",
        notes=None,
    )
    paused = TargetCompany(
        name="PausedCo_PR71",
        tier=2,
        ats="lever",
        ats_handle=None,
        notes="Paused: ATS handle unknown, soft-paused (PR #65)",
    )
    db_session.add_all([live, paused])
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/companies")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    by_name = {c["name"]: c for c in resp.json()["items"]}

    assert by_name["LiveCo_PR71"]["ats"] == "greenhouse"
    assert by_name["LiveCo_PR71"]["ats_handle"] == "livehandle"
    assert by_name["LiveCo_PR71"]["notes"] is None

    assert by_name["PausedCo_PR71"]["ats"] == "lever"
    assert by_name["PausedCo_PR71"]["ats_handle"] is None
    assert by_name["PausedCo_PR71"]["notes"] == "Paused: ATS handle unknown, soft-paused (PR #65)"


# ── /outcomes ────────────────────────────────────────────────────────────────


def _outcome(
    *,
    received_at: datetime,
    job_posting_id: uuid.UUID | None = None,
    target_company_id: uuid.UUID | None = None,
    outcome_type: str = "rejection_pre_screen",
    subject: str = "x",
    from_domain: str = "example.com",
    email_thread_id: str | None = None,
    raw_snippet: str | None = None,
) -> OutcomeEvent:
    return OutcomeEvent(
        job_posting_id=job_posting_id,
        target_company_id=target_company_id,
        email_message_id=f"msg-{uuid.uuid4().hex}",
        email_thread_id=email_thread_id,
        from_address="r@example.com",
        from_domain=from_domain,
        subject=subject,
        received_at=received_at,
        outcome_type=outcome_type,
        classifier_version="gemini-flash-lite-v1",
        classifier_confidence=0.9,
        raw_snippet=raw_snippet,
    )


@_NEEDS_DB
async def test_outcomes_chronological(db_session: Any) -> None:
    base = datetime(2026, 5, 1, tzinfo=UTC)
    db_session.add_all(
        [
            _outcome(received_at=base + timedelta(days=2)),
            _outcome(received_at=base + timedelta(days=0)),
            _outcome(received_at=base + timedelta(days=1)),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/outcomes")
    finally:
        await _drop_override()

    ts = [item["received_at"] for item in resp.json()["items"]]
    assert ts == sorted(ts)


@_NEEDS_DB
async def test_outcomes_filter_by_posting_id(db_session: Any) -> None:
    c = _company(name="OutcomesFilterCo")
    db_session.add(c)
    await db_session.flush()
    jp_a = _posting(target_company_id=c.id)
    jp_b = _posting(target_company_id=c.id)
    db_session.add_all([jp_a, jp_b])
    await db_session.flush()
    base = datetime(2026, 5, 1, tzinfo=UTC)
    db_session.add_all(
        [
            _outcome(received_at=base, job_posting_id=jp_a.id),
            _outcome(received_at=base + timedelta(hours=1), job_posting_id=jp_b.id),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/outcomes?posting_id={jp_a.id}")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    assert len(items) == 1
    assert items[0]["posting_id"] == str(jp_a.id)


@_NEEDS_DB
async def test_outcomes_response_shape(db_session: Any) -> None:
    db_session.add(
        _outcome(
            received_at=datetime(2026, 5, 1, tzinfo=UTC),
            outcome_type="application_confirmation",
        )
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/outcomes")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    assert items, "expected at least one outcome row"
    sample = items[-1]  # newest at the end of ASC order
    assert set(sample) == {
        "id",
        "posting_id",
        "target_company_id",
        "received_at",
        "stage",
        "confidence",
        "company_name",
        "subject",
        "from_domain",
        "email_thread_id",
        "raw_snippet",
        # feat/applied-unified: posting-specific overlay fields (NULL here — the
        # fixture outcome is unlinked).
        "posting_title",
        "manual_status",
    }
    assert sample["posting_title"] is None
    assert sample["manual_status"] is None
    assert sample["stage"] == "application_confirmation"
    assert sample["confidence"] == pytest.approx(0.9)


@_NEEDS_DB
async def test_outcomes_carry_label_fields_for_linked_and_unlinked(db_session: Any) -> None:
    """feat/pipeline-outcome-cards: every row must carry subject + from_domain
    (the Pipeline card label source) — for a company-LINKED row (company_name
    populated via the join) AND an UNLINKED row (company_name null, labelled
    from subject). Asserts through the endpoint serialization, not the query."""
    c = _company(name="Solv Health")
    db_session.add(c)
    await db_session.flush()
    base = datetime(2026, 5, 1, tzinfo=UTC)
    db_session.add_all(
        [
            _outcome(
                received_at=base,
                target_company_id=c.id,
                outcome_type="application_confirmation",
                subject="Thank you for applying to Solv Health",
                from_domain="greenhouse.io",
                raw_snippet="Thanks for applying to the Senior PM role at Solv Health…",
            ),
            _outcome(
                received_at=base + timedelta(hours=1),
                target_company_id=None,  # unlinked — the 119-row majority
                outcome_type="application_confirmation",
                subject="Thank you for applying to Uphold!",
                from_domain="ashbyhq.com",
            ),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/outcomes")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    by_subject = {it["subject"]: it for it in items}
    linked = by_subject["Thank you for applying to Solv Health"]
    unlinked = by_subject["Thank you for applying to Uphold!"]

    assert linked["company_name"] == "Solv Health"
    assert linked["from_domain"] == "greenhouse.io"
    # feat/pipeline-detail: the ~200-char snippet flows through for the panel.
    assert linked["raw_snippet"] == "Thanks for applying to the Senior PM role at Solv Health…"
    # Unlinked still carries subject + from_domain so the client can label it.
    assert unlinked["company_name"] is None
    assert unlinked["subject"] == "Thank you for applying to Uphold!"
    assert unlinked["from_domain"] == "ashbyhq.com"
    assert unlinked["raw_snippet"] is None


@_NEEDS_DB
async def test_outcomes_job_related_filter_excludes_noise(db_session: Any) -> None:
    """``?job_related=true`` drops unrelated/unclassified (the 1,687 noise rows
    in prod); lifecycle rows remain."""
    base = datetime(2026, 5, 1, tzinfo=UTC)
    db_session.add_all(
        [
            _outcome(received_at=base, outcome_type="application_confirmation"),
            _outcome(received_at=base + timedelta(hours=1), outcome_type="rejection_post_screen"),
            _outcome(received_at=base + timedelta(hours=2), outcome_type="unrelated"),
            _outcome(received_at=base + timedelta(hours=3), outcome_type="unclassified"),
        ]
    )
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            unfiltered = await ac.get("/outcomes")
            filtered = await ac.get("/outcomes?job_related=true")
    finally:
        await _drop_override()

    assert unfiltered.json()["total"] == 4
    fj = filtered.json()
    assert fj["total"] == 2
    kinds = {it["stage"] for it in fj["items"]}
    assert kinds == {"application_confirmation", "rejection_post_screen"}


# ── Slice 2b: best_fit_semantic blend ────────────────────────────────────────


async def _set_similarity_weight(db_session: Any, w: float) -> None:
    from sqlalchemy import select

    from job_assist.db.models import OperatorProfile

    prof = (
        await db_session.execute(select(OperatorProfile).where(OperatorProfile.id == 1))
    ).scalar_one_or_none()
    if prof is None:
        db_session.add(
            OperatorProfile(id=1, looking_for_text="", role_keywords=[], similarity_weight=w)
        )
    else:
        prof.similarity_weight = w
    await db_session.commit()


async def _seed_blend_postings(db_session: Any) -> tuple[str, str, str]:
    """A: high fit / low sim, B: low fit / high sim, C: un-embedded (sim NULL)."""
    c = _company(name="BlendCo")
    db_session.add(c)
    await db_session.flush()
    a = _posting(target_company_id=c.id, fit_score=90, similarity_score=10)
    b = _posting(target_company_id=c.id, fit_score=50, similarity_score=95)
    cc = _posting(target_company_id=c.id, fit_score=70, similarity_score=None)
    db_session.add_all([a, b, cc])
    await db_session.flush()
    db_session.add_all(
        [
            _posting_source(job_posting_id=a.id),
            _posting_source(job_posting_id=b.id),
            _posting_source(job_posting_id=cc.id),
        ]
    )
    await db_session.commit()
    return str(a.id), str(b.id), str(cc.id)


def _ids(resp: Any) -> list[str]:
    return [it["id"] for it in resp.json()["items"]]


async def _order(db_session: Any, sort: str) -> list[str]:
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(f"/postings?sort={sort}&per_company_cap=0")
    finally:
        await _drop_override()
    return _ids(resp)


@_NEEDS_DB
async def test_best_fit_semantic_w0_byte_identical_to_best_fit(db_session: Any) -> None:
    await _set_similarity_weight(db_session, 0.0)
    a, b, cc = await _seed_blend_postings(db_session)
    bf = await _order(db_session, "best_fit")
    bfs = await _order(db_session, "best_fit_semantic")
    assert bf == bfs  # at w=0 the blend collapses to fit_score
    assert bf == [a, cc, b]  # fit DESC: 90, 70, 50


@_NEEDS_DB
async def test_best_fit_semantic_w1_high_sim_low_fit_rises(db_session: Any) -> None:
    await _set_similarity_weight(db_session, 1.0)
    a, b, cc = await _seed_blend_postings(db_session)
    # w=1 → order by COALESCE(sim, fit): B=95, C=70(fallback), A=10
    assert await _order(db_session, "best_fit_semantic") == [b, cc, a]


@_NEEDS_DB
async def test_best_fit_semantic_monotonic_in_w(db_session: Any) -> None:
    _a, b, _cc = await _seed_blend_postings(db_session)
    positions = []
    for w in (0.0, 0.5, 1.0):
        await _set_similarity_weight(db_session, w)
        positions.append((await _order(db_session, "best_fit_semantic")).index(b))
    # B (low fit, high sim) rises monotonically as w grows: last → first.
    assert positions[0] > positions[1] >= positions[2]


@_NEEDS_DB
async def test_best_fit_semantic_unembedded_falls_back_to_fit(db_session: Any) -> None:
    await _set_similarity_weight(db_session, 1.0)
    _a, _b, cc = await _seed_blend_postings(db_session)
    # C (sim NULL) ranks by its fit_score (70) — NOT dropped to NULLS LAST.
    assert (await _order(db_session, "best_fit_semantic")).index(cc) == 1


@_NEEDS_DB
async def test_best_fit_semantic_reversible(db_session: Any) -> None:
    await _seed_blend_postings(db_session)
    await _set_similarity_weight(db_session, 0.0)
    o_before = await _order(db_session, "best_fit_semantic")
    await _set_similarity_weight(db_session, 0.5)
    await _order(db_session, "best_fit_semantic")
    await _set_similarity_weight(db_session, 0.0)
    o_after = await _order(db_session, "best_fit_semantic")
    assert o_before == o_after  # 0 → 0.5 → 0 returns to the identical order


@_NEEDS_DB
async def test_postings_surfaces_similarity_score(db_session: Any) -> None:
    a, _b, cc = await _seed_blend_postings(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=0")
    finally:
        await _drop_override()
    by_id = {it["id"]: it for it in resp.json()["items"]}
    assert by_id[a]["similarity_score"] == 10
    assert by_id[cc]["similarity_score"] is None  # un-embedded surfaces as null


# ── triage export / list parity (fix/triage-export-state-lateral) ────────────


def _export_job_titles(xlsx_bytes: bytes) -> list[str]:
    """Return the raw_title column values from the export's 'Jobs' sheet.

    Reads through the actual xlsx the endpoint produced — the bug lived in
    the export's SQL row query, so the test must parse the export output,
    not re-run a list query.
    """
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    jobs = wb["Jobs"]
    rows = list(jobs.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    # The Jobs sheet has a 'Role' column carrying the posting's raw_title.
    role_idx = next((i for i, h in enumerate(header) if h.strip().lower() == "role"), None)
    if role_idx is None:
        # Fall back to scanning every cell so the test still asserts presence.
        return [str(c) for row in rows[1:] for c in row if c is not None]
    return [str(row[role_idx]) for row in rows[1:] if row[role_idx] is not None]


@_NEEDS_DB
async def test_triage_export_includes_unactioned_rows(db_session: Any) -> None:
    """Regression: a ``state=triage`` export must return the un-actioned
    postings, not 0.

    The export's row query references the ``recent_pa`` state LATERAL via the
    triage WHERE predicate. If that lateral isn't OUTER-joined onto the row
    SELECT, it folds in as an implicit INNER lateral that drops every
    un-actioned posting (the lateral yields no row when no action exists) —
    exactly the rows ``triage`` (``pa_action_type IS NULL``) selects. The
    export then came back empty while ``list_postings`` (which outer-joins)
    returned the same rows. Exercise the EXPORT path specifically.
    """
    from job_assist.db.models import PostingAction

    company = _company(name="ExportCo", tier=1)
    db_session.add(company)
    await db_session.flush()

    # Two un-actioned (triage) postings + one acted-on (leaves triage).
    triage_a = _posting(target_company_id=company.id, title="Triage Role A", fit_score=90)
    triage_b = _posting(target_company_id=company.id, title="Triage Role B", fit_score=80)
    actioned = _posting(target_company_id=company.id, title="Actioned Role", fit_score=70)
    for jp in (triage_a, triage_b, actioned):
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    db_session.add(PostingAction(job_posting_id=actioned.id, action_type="interested"))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            # cap off so both same-company triage rows survive the per-company cap.
            resp = await ac.get("/postings/export.xlsx?state=triage&per_company_cap=0")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    titles = _export_job_titles(resp.content)
    # The pre-fix bug returned ZERO rows here.
    assert len(titles) == 2, f"expected 2 un-actioned triage rows, got {len(titles)}: {titles}"
    # The Jobs sheet emits normalized_title (lowercased) — match accordingly.
    joined = " ".join(titles).lower()
    assert "triage role a" in joined
    assert "triage role b" in joined
    # The acted-on posting must NOT leak into the triage export.
    assert "actioned role" not in joined


@_NEEDS_DB
async def test_actioned_state_export_still_filters(db_session: Any) -> None:
    """Control for the triage fix: a non-triage state export (``interested``)
    returns exactly the acted-on rows and excludes un-actioned ones — proving
    the added outer-join didn't loosen the state filter into a pass-through.
    """
    from job_assist.db.models import PostingAction

    company = _company(name="ActionedCo", tier=1)
    db_session.add(company)
    await db_session.flush()

    triage_only = _posting(target_company_id=company.id, title="Untouched Role", fit_score=90)
    interested = _posting(target_company_id=company.id, title="Interested Role", fit_score=80)
    for jp in (triage_only, interested):
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    db_session.add(PostingAction(job_posting_id=interested.id, action_type="interested"))
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings/export.xlsx?state=interested&per_company_cap=0")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    titles = _export_job_titles(resp.content)
    joined = " ".join(titles).lower()
    assert "interested role" in joined
    assert "untouched role" not in joined


# ── triage export = current filtered view, unbounded (feat/triage-export-full-view) ──


@_NEEDS_DB
async def test_export_matches_filtered_sorted_list_exactly(db_session: Any) -> None:
    """The export is the list view minus pagination: SAME rows, SAME order.

    Drive both ``GET /postings`` and ``GET /postings/export.xlsx`` with the
    identical filter+sort URL, then assert the export's ordered titles equal the
    list's ordered titles. Both run through ``build_view_parts(spec)``, so this
    pins them together — if either drifts (a filter, the sort, the cap), the
    ordered sequences diverge and this fails.
    """
    company = _company(name="ParityCo", tier=1)
    db_session.add(company)
    await db_session.flush()

    # Distinct fit_scores → an unambiguous best_fit (score DESC) order.
    scored = [("Parity Role Alpha", 95), ("Parity Role Bravo", 81), ("Parity Role Charlie", 73)]
    for title, score in scored:
        jp = _posting(target_company_id=company.id, title=title, fit_score=score)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    # Identical params on both endpoints. cap off + high limit so the list
    # returns the FULL set (the only difference left is export's missing limit).
    query = "state=triage&per_company_cap=0&sort=best_fit"
    ac = await _client(db_session)
    try:
        async with ac:
            list_resp = await ac.get(f"/postings?{query}&limit=100")
            export_resp = await ac.get(f"/postings/export.xlsx?{query}")
    finally:
        await _drop_override()

    assert list_resp.status_code == 200
    assert export_resp.status_code == 200
    list_titles = [item["role"]["title"] for item in list_resp.json()["items"]]
    export_titles = _export_job_titles(export_resp.content)
    # Same rows, same order — the export IS the list, unpaginated.
    assert export_titles == list_titles
    assert len(export_titles) == 3


@_NEEDS_DB
async def test_export_has_no_row_cap(db_session: Any) -> None:
    """The hardcoded 40-row cap is gone: a filtered view of >40 rows exports
    ALL of them, not the old top-40 slice."""
    company = _company(name="BigCo", tier=1)
    db_session.add(company)
    await db_session.flush()

    n = 45  # > the old EXPORT_ROW_CAP of 40
    for i in range(n):
        jp = _posting(
            target_company_id=company.id,
            title=f"Capless Role {i:02d}",
            fit_score=90 - i,
        )
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    # per_company_cap=0 so all 45 same-company rows survive the per-company cap;
    # the only thing that could clamp to 40 was the removed export limit.
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings/export.xlsx?state=triage&per_company_cap=0")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    titles = _export_job_titles(resp.content)
    assert len(titles) == n, f"expected all {n} rows (cap gone), got {len(titles)}"


@_NEEDS_DB
async def test_export_empty_filter_yields_headers_only_not_error(db_session: Any) -> None:
    """A filter matching 0 rows downloads a valid xlsx with headers only —
    never a 4xx/5xx."""
    company = _company(name="EmptyCo", tier=1)
    db_session.add(company)
    await db_session.flush()
    jp = _posting(target_company_id=company.id, title="Only Role", fit_score=88)
    db_session.add(jp)
    await db_session.flush()
    db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()

    # target_company_id pointing at a company with no postings → 0 matches.
    empty_company = _company(name="NoPostingsCo", tier=1)
    db_session.add(empty_company)
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get(
                f"/postings/export.xlsx?state=triage&per_company_cap=0&target_company_id={empty_company.id}"
            )
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Valid workbook, headers intact, zero data rows.
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    assert wb.sheetnames == ["Export Context", "Jobs"]
    jobs_rows = list(wb["Jobs"].iter_rows(values_only=True))
    assert jobs_rows[0][0] == "rank"  # header present
    assert _export_job_titles(resp.content) == []  # no data rows

"""Tests for ``GET /postings?per_company_cap=N`` (PR #58).

The cap is the operator-facing version of "don't drown me in 14
postings from the same Workday tenant." Implementation is a ROW_NUMBER
CTE partitioned by ``COALESCE(target_company_id::text, id::text)`` so
each NULL-company row is its own bucket (exempt from the cap).

These tests pin:
  - cap=N reduces the row count correctly across mixed-company fixtures
  - cap=0 disables the cap entirely (back-compat)
  - NULL ``target_company_id`` rows are exempt
  - Ranking inside a bucket is fixed at ``fit_score DESC NULLS LAST,
    first_seen_at DESC, id ASC`` regardless of the operator's outer sort
  - The outer sort orders the surviving rows
  - Pagination math matches the capped count, not the uncapped count
  - 2-query budget preserved (CTE inlines, no extra round-trip)
  - Stable ``id ASC`` tiebreaker inside the bucket
  - ``per_company_cap=-1`` → 422
"""

from __future__ import annotations

import os
import uuid
from datetime import UTC, datetime, timedelta
from typing import Any

import pytest
from httpx import ASGITransport, AsyncClient

from job_assist.db.models import JobPosting, PostingSource, TargetCompany

_NEEDS_DB = pytest.mark.skipif(
    not os.getenv("TEST_DATABASE_URL"),
    reason="TEST_DATABASE_URL not set",
)


# ── Client + execute counter (mirror existing pattern) ─────────────────────


async def _client(db_session: Any) -> AsyncClient:
    from job_assist.db.session import get_db
    from job_assist.main import app

    async def _override() -> Any:
        yield db_session

    app.dependency_overrides[get_db] = _override
    return AsyncClient(transport=ASGITransport(app=app), base_url="http://test")


async def _drop_override() -> None:
    from job_assist.db.session import get_db
    from job_assist.main import app

    app.dependency_overrides.pop(get_db, None)


class _ExecuteCounter:
    def __init__(self, session: Any) -> None:
        self._session = session
        self._original = session.execute
        self.count = 0

    async def _wrapped(self, *args: Any, **kwargs: Any) -> Any:
        self.count += 1
        return await self._original(*args, **kwargs)

    def __enter__(self) -> _ExecuteCounter:
        self._session.execute = self._wrapped  # type: ignore[method-assign]
        return self

    def __exit__(self, *_exc: Any) -> None:
        self._session.execute = self._original  # type: ignore[method-assign]

    async def __aenter__(self) -> _ExecuteCounter:
        return self.__enter__()

    async def __aexit__(self, *_exc: Any) -> None:
        self.__exit__(*_exc)


# ── Validation (pure) ──────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_negative_cap_returns_422() -> None:
    from job_assist.main import app

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as ac:
        resp = await ac.get("/postings?per_company_cap=-1")
    assert resp.status_code == 422


# ── Fixture factories ──────────────────────────────────────────────────────


def _company(name: str, tier: int = 1) -> TargetCompany:
    return TargetCompany(
        name=name,
        tier=tier,
        ats="greenhouse",
        ats_handle=f"handle-{uuid.uuid4().hex[:6]}",
    )


def _posting(
    *,
    target_company_id: uuid.UUID | None,
    first_seen_at: datetime,
    fit_score: int | None = None,
    title: str = "Senior PM",
) -> JobPosting:
    suffix = uuid.uuid4().hex[:10]
    return JobPosting(
        canonical_company_name="TestCo",
        target_company_id=target_company_id,
        normalized_title=title.lower(),
        raw_title=title,
        jd_text="JD body.",
        jd_text_hash=f"{'0' * 54}{suffix}",
        content_hash=f"hash-{suffix}",
        first_seen_at=first_seen_at,
        last_seen_at=first_seen_at,
        fit_score=fit_score,
    )


def _posting_source(*, job_posting_id: uuid.UUID) -> PostingSource:
    """Mirror the canonical factory shape — all 8 NOT NULL columns."""
    return PostingSource(
        job_posting_id=job_posting_id,
        ats="greenhouse",
        source_job_id=uuid.uuid4().hex,
        source_url=f"https://example.test/{uuid.uuid4().hex[:8]}",
        apply_url=None,
        raw_payload={},
        parser_version="test-v1",
        fetch_status="ok",
        fetched_at=datetime.now(tz=UTC),
    )


async def _seed_cap_fixture(db_session: Any) -> dict[str, list[uuid.UUID]]:
    """Brief-specified fixture:
      - 5 postings from company A (fit_score 100, 90, 80, 70, 60)
      - 2 postings from company B (fit_score 95, 75)
      - 3 postings with NULL target_company_id (fit_score 50, 40, 30)
    Total: 10 rows.

    Each posting's first_seen_at is staggered so ordering is unambiguous.
    """
    now = datetime.now(tz=UTC)

    co_a = _company("CompA", tier=1)
    co_b = _company("CompB", tier=2)
    db_session.add_all([co_a, co_b])
    await db_session.flush()

    # company A: 5 postings, scores 100→60, oldest-first to newest
    a_ids: list[uuid.UUID] = []
    for idx, score in enumerate([100, 90, 80, 70, 60]):
        jp = _posting(
            target_company_id=co_a.id,
            first_seen_at=now - timedelta(hours=idx),
            fit_score=score,
        )
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
        a_ids.append(jp.id)

    b_ids: list[uuid.UUID] = []
    for idx, score in enumerate([95, 75]):
        jp = _posting(
            target_company_id=co_b.id,
            first_seen_at=now - timedelta(hours=10 + idx),
            fit_score=score,
        )
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
        b_ids.append(jp.id)

    null_ids: list[uuid.UUID] = []
    for idx, score in enumerate([50, 40, 30]):
        jp = _posting(
            target_company_id=None,
            first_seen_at=now - timedelta(hours=20 + idx),
            fit_score=score,
        )
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
        null_ids.append(jp.id)

    await db_session.commit()
    return {"a": a_ids, "b": b_ids, "null": null_ids}


# ── Behaviour ──────────────────────────────────────────────────────────────


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_3_returns_3_from_a_2_from_b_3_null(db_session: Any) -> None:
    """cap=3: company A → top 3 by score (100, 90, 80); B → both (2);
    NULL bucket → all 3 (exempt). Total 8 rows."""
    await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=3&limit=100&sort=best_fit")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 8
    assert len(body["items"]) == 8


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_1_returns_top_1_per_bucket(db_session: Any) -> None:
    """cap=1: company A → 1 (score 100); B → 1 (score 95); NULL → 3 (exempt).
    Total 5 rows."""
    ids = await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=1&limit=100&sort=best_fit")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 5
    surfaced = [item["id"] for item in body["items"]]
    # A's top row (score 100) — first in the a_ids list because we iterate
    # 100 → 60. B's top row similarly.
    assert str(ids["a"][0]) in surfaced
    assert str(ids["b"][0]) in surfaced
    # All 3 NULL-bucket rows surface (exempt from the cap).
    for null_id in ids["null"]:
        assert str(null_id) in surfaced


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_0_disables_returns_all_10(db_session: Any) -> None:
    """cap=0: no filter, full 10 rows."""
    await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=0&limit=100")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 10
    assert len(body["items"]) == 10


@_NEEDS_DB
@pytest.mark.asyncio
async def test_ranking_inside_bucket_is_score_first_regardless_of_outer_sort(
    db_session: Any,
) -> None:
    """Ranking inside each company bucket is FIXED:
       fit_score DESC NULLS LAST, first_seen_at DESC, id ASC.
    The outer sort then orders the surviving rows.

    Brief: ``sort=oldest&per_company_cap=3`` surfaces "oldest of each
    company's top-3 by score", NOT "oldest 3 per company."
    """
    ids = await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            # outer sort=oldest, but ranking is still score-first.
            # A's top 3 by score = scores 100, 90, 80 → indices 0,1,2 in a_ids.
            # Those rows were inserted with first_seen_at offsets 0h, 1h, 2h ago.
            # So sort=oldest within A's top 3 → index 2 (80, 2h ago) first.
            resp = await ac.get("/postings?per_company_cap=3&sort=oldest&limit=100")
    finally:
        await _drop_override()

    items = resp.json()["items"]
    # The point: A's bottom 2 (scores 70, 60 → indices 3,4) MUST be absent
    # — they didn't make A's top-3 by score, so outer sort=oldest can't
    # surface them either.
    a_id_set = {str(x) for x in ids["a"]}
    surfaced_a_ids = {item["id"] for item in items if item["id"] in a_id_set}
    assert surfaced_a_ids == {str(ids["a"][0]), str(ids["a"][1]), str(ids["a"][2])}


@_NEEDS_DB
@pytest.mark.asyncio
async def test_pagination_math_matches_capped_count(db_session: Any) -> None:
    """cap=3, limit=5 → page1=5, page2=3, total=8 (NOT 10)."""
    await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            p1 = await ac.get("/postings?per_company_cap=3&limit=5&offset=0&sort=best_fit")
            p2 = await ac.get("/postings?per_company_cap=3&limit=5&offset=5&sort=best_fit")
    finally:
        await _drop_override()

    p1_body = p1.json()
    p2_body = p2.json()
    assert p1_body["total"] == 8  # capped, not 10
    assert len(p1_body["items"]) == 5
    assert p2_body["total"] == 8
    assert len(p2_body["items"]) == 3
    page_1_ids = {x["id"] for x in p1_body["items"]}
    page_2_ids = {x["id"] for x in p2_body["items"]}
    assert page_1_ids.isdisjoint(page_2_ids)


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_preserves_two_query_budget(db_session: Any) -> None:
    """The CTE inlines into both COUNT and SELECT — still ≤2 queries."""
    await _seed_cap_fixture(db_session)
    counter = _ExecuteCounter(db_session)
    ac = await _client(db_session)
    try:
        async with ac, counter:
            resp = await ac.get("/postings?per_company_cap=3&limit=100&sort=best_fit")
    finally:
        await _drop_override()

    assert resp.status_code == 200
    assert counter.count <= 2, f"cap=3 emitted {counter.count} queries (expected ≤2)"


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_with_id_tiebreaker_on_same_score_and_first_seen(
    db_session: Any,
) -> None:
    """Same (score, first_seen_at) → ordering by id ASC inside the bucket
    (bestiary lock)."""
    now = datetime.now(tz=UTC).replace(microsecond=0)
    co = _company("TieCo", tier=1)
    db_session.add(co)
    await db_session.flush()

    # 5 postings, all identical fit_score=80, identical first_seen_at —
    # ranking must fall through to id ASC.
    ids: list[uuid.UUID] = []
    for _ in range(5):
        jp = _posting(target_company_id=co.id, first_seen_at=now, fit_score=80)
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
        ids.append(jp.id)
    await db_session.commit()

    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=3&limit=100&sort=best_fit")
    finally:
        await _drop_override()

    surfaced = [item["id"] for item in resp.json()["items"]]
    # The cap picks 3 by id ASC. The remaining 2 ids must NOT appear.
    expected = [str(x) for x in sorted(ids)[:3]]
    # Don't assert exact order in the outer list (sort=best_fit + same score
    # falls through to id ASC at the outer layer too, but the test focus is
    # the bucket-level selection). Just assert membership.
    assert set(surfaced) == set(expected)


@_NEEDS_DB
@pytest.mark.asyncio
async def test_default_cap_3_applied_when_param_omitted(db_session: Any) -> None:
    """Brief: default ``per_company_cap=3``. Calling /postings without
    the param applies the cap silently."""
    ids = await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?limit=100&sort=best_fit")
    finally:
        await _drop_override()

    body = resp.json()
    # 3 from A + 2 from B + 3 NULL = 8 (same as explicit cap=3)
    assert body["total"] == 8
    # A's bottom 2 (scores 70, 60) absent
    surfaced = {item["id"] for item in body["items"]}
    assert str(ids["a"][3]) not in surfaced  # score 70
    assert str(ids["a"][4]) not in surfaced  # score 60


# ── feat/tunable-per-company-cap: the reachability fix ──────────────────────
# The cap mechanism is already covered above. These pin the NEW wiring: that
# the operator can MOVE the cap (raise it / disable it) and that moving it
# changes what surfaces — including via the persisted operator_profile default
# (the gap that left the operator stuck at 3). Per the verification standard.


@_NEEDS_DB
@pytest.mark.asyncio
async def test_cap_4_surfaces_more_than_cap_3(db_session: Any) -> None:
    """Intermediate raise: cap=4 on the 5-role company A surfaces its 4th-best
    role (score 70), which cap=3 hides. Proves raising the cap surfaces more.
    Total: 4 (A) + 2 (B) + 3 (NULL) = 9."""
    ids = await _seed_cap_fixture(db_session)
    ac = await _client(db_session)
    try:
        async with ac:
            resp = await ac.get("/postings?per_company_cap=4&limit=100&sort=best_fit")
    finally:
        await _drop_override()

    body = resp.json()
    assert body["total"] == 9
    surfaced = {item["id"] for item in body["items"]}
    assert str(ids["a"][3]) in surfaced  # score 70 NOW surfaces (hidden at cap=3)
    assert str(ids["a"][4]) not in surfaced  # score 60 still capped out


async def _set_profile_cap(db_session: Any, cap: int) -> None:
    """Set operator_profile.per_company_cap on the singleton (id=1)."""
    from sqlalchemy import select

    from job_assist.db.models import OperatorProfile

    prof = (
        await db_session.execute(select(OperatorProfile).where(OperatorProfile.id == 1))
    ).scalar_one()
    prof.per_company_cap = cap
    await db_session.commit()


async def _seed_one_company(db_session: Any, n: int) -> uuid.UUID:
    """One company with ``n`` postings, fit_scores 100, 99, …; staggered
    first_seen so ordering is deterministic. Returns the company id."""
    now = datetime.now(tz=UTC)
    co = _company("CapCo", tier=1)
    db_session.add(co)
    await db_session.flush()
    for idx in range(n):
        jp = _posting(
            target_company_id=co.id,
            first_seen_at=now - timedelta(hours=idx),
            fit_score=100 - idx,
        )
        db_session.add(jp)
        await db_session.flush()
        db_session.add(_posting_source(job_posting_id=jp.id))
    await db_session.commit()
    return co.id


@_NEEDS_DB
@pytest.mark.asyncio
async def test_endpoint_reads_profile_cap_default_and_param_overrides(db_session: Any) -> None:
    """The reachability proof. Company with 6 roles:
    - profile cap=6, no param  → 6 surface (operator default honored)
    - ?per_company_cap=2       → 2 surface (explicit override wins)
    - profile cap=0, no param  → all 6 surface (disabled)
    """
    await _seed_one_company(db_session, 6)

    # profile default = 6 → no-param view surfaces all 6
    await _set_profile_cap(db_session, 6)
    ac = await _client(db_session)
    try:
        async with ac:
            no_param = await ac.get("/postings?limit=100&sort=best_fit")
            override = await ac.get("/postings?per_company_cap=2&limit=100&sort=best_fit")
        assert no_param.json()["total"] == 6, "profile cap=6 must surface all 6 (operator default)"
        assert override.json()["total"] == 2, (
            "explicit ?per_company_cap=2 must override the profile"
        )

        # profile cap=0 → disabled → all surface with no param
        await _set_profile_cap(db_session, 0)
        async with await _client(db_session):
            from job_assist.db.session import get_db
            from job_assist.main import app

            async def _ov() -> Any:
                yield db_session

            app.dependency_overrides[get_db] = _ov
            ac2 = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
            async with ac2:
                disabled = await ac2.get("/postings?limit=100&sort=best_fit")
        assert disabled.json()["total"] == 6, "profile cap=0 must disable the cap (show all)"
    finally:
        await _set_profile_cap(db_session, 3)  # restore singleton default
        await _drop_override()


@_NEEDS_DB
@pytest.mark.asyncio
async def test_export_honors_profile_cap_matching_the_view(db_session: Any) -> None:
    """Export parity: with profile cap=2 and a 6-role company, both the list
    view AND the xlsx export surface exactly 2 rows — the "exported == visible"
    contract holds through the profile-default fallback."""
    from io import BytesIO

    from openpyxl import load_workbook

    await _seed_one_company(db_session, 6)
    await _set_profile_cap(db_session, 2)
    ac = await _client(db_session)
    try:
        async with ac:
            view = await ac.get("/postings?limit=100&sort=best_fit")
            export = await ac.get("/postings/export.xlsx?sort=best_fit")
    finally:
        await _set_profile_cap(db_session, 3)
        await _drop_override()

    view_total = view.json()["total"]
    assert view_total == 2  # profile cap applied to the view

    wb = load_workbook(BytesIO(export.content))
    ws = wb["Jobs"]
    # Column 1 is the integer rank; count data rows (skip the header).
    data_rows = [r[0].value for r in ws.iter_rows(min_row=2) if isinstance(r[0].value, int)]
    assert len(data_rows) == view_total, "export row count must match the capped view"

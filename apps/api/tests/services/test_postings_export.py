"""Unit tests for services/postings_export.py (feat/triage-export-xlsx).

Pure tests — no DB. Build a few JobPosting+TargetCompany rows in memory,
run them through ``build_workbook_bytes``, then parse the result back
with openpyxl to confirm structure (two sheets, headers, rank ordering,
score breakdown columns).
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from job_assist.db.enums import RoleFamily, SeniorityLevel
from job_assist.db.models.job_posting import JobPosting
from job_assist.db.models.operator_profile import OperatorProfile
from job_assist.db.models.target_company import TargetCompany
from job_assist.services.postings_export import build_workbook_bytes
from job_assist.services.postings_query import PostingsViewSpec


def _posting(**overrides: Any) -> JobPosting:
    now = datetime.now(tz=UTC)
    suffix = uuid.uuid4().hex[:10]
    defaults: dict[str, Any] = {
        "canonical_company_name": "TestCo",
        "target_company_id": None,
        "normalized_title": "senior product manager",
        "raw_title": "Senior Product Manager",
        "jd_text": "JD body.",
        "jd_text_hash": f"{'0' * 54}{suffix}",
        "content_hash": f"hash-{suffix}",
        "first_seen_at": now,
        "last_seen_at": now,
        "role_family": RoleFamily.product_management.value,
        "seniority_level": SeniorityLevel.senior_pm.value,
        "remote_type": "remote",
        "salary_min": 150_000,
        "salary_max": 200_000,
        "salary_currency": "USD",
        "salary_period": "annual",
        "locations_normalized": [{"remote_type": "remote", "city": "Remote"}],
        "fit_score": 92,
        "jd_summary_markdown": "## Role\n\n- Build stuff.",
    }
    defaults.update(overrides)
    return JobPosting(**defaults)


def _target_company(name: str = "Acme", tier: int | None = 1) -> TargetCompany:
    return TargetCompany(
        id=uuid.uuid4(),
        name=name,
        domain=f"{name.lower()}.com",
        tier=tier,
    )


def _profile() -> OperatorProfile:
    now = datetime.now(tz=UTC)
    return OperatorProfile(
        id=1,
        looking_for_text="PM roles",
        role_keywords=[],
        geo_whitelist=["Remote", "Austin"],
        salary_floor_usd=100_000,
        salary_ceiling_usd=250_000,
        applicant_cap=500,
        seniority_levels_included=["senior_pm", "lead_pm"],
        staffing_firm_blocklist=[],
        created_at=now,
        updated_at=now,
    )


def _open(buf: bytes) -> Any:
    return load_workbook(BytesIO(buf))


@pytest.fixture
def sample_rows() -> list[tuple[JobPosting, TargetCompany | None, str | None, str | None]]:
    return [
        (
            _posting(fit_score=92),
            _target_company("Acme", tier=1),
            "greenhouse",
            "https://a.example/1",
        ),
        (_posting(fit_score=70), _target_company("Beta", tier=2), "lever", "https://b.example/2"),
        (_posting(fit_score=None, canonical_company_name="Gamma"), None, None, None),
    ]


def test_workbook_renders_all_rows_no_cap() -> None:
    """The 40-row cap is gone: every row handed in is rendered (the endpoint
    no longer clamps the SQL fetch, so the workbook must not clamp either)."""
    rows = [
        (_posting(fit_score=90 - i), _target_company(f"Co{i}", tier=1), "greenhouse", f"https://x/{i}")
        for i in range(57)  # well past the old 40 cap
    ]
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=rows,
        corpus_size=999,
        matched_before_cap=57,
    )
    ws = _open(buf)["Jobs"]
    assert ws.max_row == 58  # 1 header + 57 data rows
    # Ranks run 1..57 contiguously — nothing truncated.
    assert [ws.cell(row=i, column=1).value for i in range(2, 59)] == list(range(1, 58))


def test_empty_rows_yields_headers_only_file_not_error() -> None:
    """0 matches → a VALID workbook with header rows only (edge case: an empty
    filtered view must download a file, never error)."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(tier=(4,)),
        profile=_profile(),
        rows=[],
        corpus_size=999,
        matched_before_cap=0,
    )
    wb = _open(buf)
    assert wb.sheetnames == ["Export Context", "Jobs"]
    jobs = wb["Jobs"]
    assert jobs.max_row == 1  # header row only, no data
    assert jobs.cell(row=1, column=1).value == "rank"  # headers intact


def test_context_sheet_advertises_no_row_cap() -> None:
    """Sheet 1's 'Export row cap' line must say there's no cap (not '40')."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=[(_posting(), _target_company(), "greenhouse", "https://x")],
        corpus_size=10,
        matched_before_cap=1,
    )
    ws = _open(buf)["Export Context"]
    # Find the 'Export row cap' label cell (col A) and read its value (col B).
    cap_value = next(
        (
            ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "Export row cap"
        ),
        None,
    )
    assert cap_value is not None
    assert "full filtered view" in cap_value
    assert "40" not in cap_value


def test_workbook_has_two_sheets_in_correct_order(
    sample_rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
) -> None:
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=sample_rows,
        corpus_size=999,
        matched_before_cap=142,
    )
    wb = _open(buf)
    # Context first so reviewer sees provenance on open.
    assert wb.sheetnames == ["Export Context", "Jobs"]


def test_jobs_sheet_headers_present(
    sample_rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
) -> None:
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=sample_rows,
        corpus_size=10,
        matched_before_cap=3,
    )
    wb = _open(buf)
    ws = wb["Jobs"]
    headers = [c.value for c in ws[1]]
    # Spot-check the load-bearing columns; locking the full list would
    # make column re-orderings noisy without adding signal.
    for required in (
        "rank",
        "company",
        "role",
        "fit_score",
        "score.role_family",
        "score.seniority",
        "score.salary",
        "score.tier",
        "score.geo",
        "apply_url",
        "jd_summary_markdown",
    ):
        assert required in headers


def test_jobs_sheet_rows_match_input_count_and_rank(
    sample_rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
) -> None:
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=sample_rows,
        corpus_size=10,
        matched_before_cap=3,
    )
    wb = _open(buf)
    ws = wb["Jobs"]
    # 1 header + 3 data rows
    assert ws.max_row == 4
    ranks = [ws.cell(row=i, column=1).value for i in range(2, 5)]
    assert ranks == [1, 2, 3]


def test_jobs_sheet_carries_score_breakdown(
    sample_rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
) -> None:
    """Score breakdown is computed on the fly; integers 0-100 each."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=sample_rows,
        corpus_size=10,
        matched_before_cap=3,
    )
    wb = _open(buf)
    ws = wb["Jobs"]
    headers = [c.value for c in ws[1]]
    for bucket in ("role_family", "seniority", "salary", "tier", "geo"):
        col_idx = headers.index(f"score.{bucket}") + 1
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            assert isinstance(v, int)
            assert 0 <= v <= 100


def test_context_sheet_reports_counts_and_filters() -> None:
    spec = PostingsViewSpec(
        tier=(1, 2),
        ats=("greenhouse",),
        sort="best_fit",
        per_company_cap=3,
        include_closed=True,
    )
    buf = build_workbook_bytes(
        spec=spec,
        profile=_profile(),
        rows=[(_posting(fit_score=88), _target_company(), "greenhouse", "https://x")],
        corpus_size=1234,
        matched_before_cap=87,
    )
    wb = _open(buf)
    ws = wb["Export Context"]
    # Flatten all cell values into one string for substring assertions.
    blob = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "1234" in blob  # corpus_size
    assert "87" in blob  # matched_before_cap
    assert "best_fit" in blob  # sort
    assert "greenhouse" in blob  # ats filter
    assert "yes" in blob  # include_closed=True formatted as "yes"


def test_context_sheet_includes_scorer_weights_and_hard_rules() -> None:
    """Scorer weights + hard-rule thresholds give the reviewing AI calibration."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=[(_posting(), _target_company(), "greenhouse", "https://x")],
        corpus_size=10,
        matched_before_cap=5,
    )
    wb = _open(buf)
    blob = "\n".join(
        str(c.value) for row in wb["Export Context"].iter_rows() for c in row if c.value is not None
    )
    # Weights surface for each of the five buckets.
    for bucket in ("role_family", "seniority", "salary", "tier", "geo"):
        assert bucket in blob
    # Hard-rule context surfaces.
    assert "salary_floor_usd" in blob
    assert "geo_whitelist" in blob
    assert "applicant_cap" in blob


def test_context_sheet_score_range_handles_all_null_fit_scores() -> None:
    """When every row has fit_score=None, the range falls back to a marker."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=[(_posting(fit_score=None), _target_company(), None, None)],
        corpus_size=1,
        matched_before_cap=1,
    )
    wb = _open(buf)
    blob = "\n".join(
        str(c.value) for row in wb["Export Context"].iter_rows() for c in row if c.value is not None
    )
    assert "no scored rows" in blob


def test_unmatched_posting_uses_canonical_company_name() -> None:
    """When target_company is None, the row should still carry a company name."""
    buf = build_workbook_bytes(
        spec=PostingsViewSpec(),
        profile=_profile(),
        rows=[(_posting(canonical_company_name="Solo Inc"), None, None, None)],
        corpus_size=1,
        matched_before_cap=1,
    )
    wb = _open(buf)
    ws = wb["Jobs"]
    company_col = [c.value for c in ws[1]].index("company") + 1
    assert ws.cell(row=2, column=company_col).value == "Solo Inc"

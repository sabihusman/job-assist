"""xlsx export of the Triage view (feat/triage-export-xlsx).

Builds a two-sheet workbook from the same query parts ``GET /postings``
uses, so the exported rows are exactly the visible rows for a given URL.
Sheet 1 carries enough provenance (timestamp, filters, sort, thresholds,
weights, score range) that an external reviewing AI can interpret the
slice; sheet 2 carries the rows themselves with the score breakdown
computed on the fly.

Pure-ish: takes the already-fetched row tuples + the resolved profile +
the matched-before-cap count and returns the workbook bytes. The endpoint
in main.py handles the SQL execution and HTTP shape.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from job_assist.db.models import JobPosting, OperatorProfile, TargetCompany
from job_assist.services.postings_query import PostingsViewSpec
from job_assist.services.scoring import (
    _WEIGHTS,
    PREFERRED_FAMILIES,
    SCORER_VERSION,
    score_breakdown,
)
from job_assist.triage.config import hard_rule_config_from_profile

# Max rows on Sheet 2. Locked to 40 (PR-scope decision): "top 40 by the
# visible sort." The endpoint clamps the ``limit`` query param to this
# value before the SQL fetch — see ``main.py::export_postings_xlsx``.
EXPORT_ROW_CAP = 40

_HEADER_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")


def _enum_value(v: Any) -> str | None:
    if v is None:
        return None
    inner = getattr(v, "value", v)
    return str(inner) if inner is not None else None


def _fmt_filter_list(values: tuple[Any, ...]) -> str:
    return ", ".join(str(v) for v in values) if values else "(none)"


def _fmt_bool(v: bool) -> str:
    return "yes" if v else "no"


def _build_context_sheet(
    ws: Any,
    *,
    spec: PostingsViewSpec,
    profile: OperatorProfile,
    corpus_size: int,
    matched_before_cap: int,
    visible_rows: int,
    score_min: int | None,
    score_max: int | None,
) -> None:
    """Sheet 1: provenance + thresholds + plain-language notes."""
    ws.title = "Export Context"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80

    hard = hard_rule_config_from_profile(profile)
    now = datetime.now(UTC).replace(microsecond=0).isoformat()

    sections: list[tuple[str, list[tuple[str, str]]]] = [
        (
            "Export",
            [
                ("Generated (UTC)", now),
                ("Scorer version", SCORER_VERSION),
                ("Export row cap", str(EXPORT_ROW_CAP)),
            ],
        ),
        (
            "Counts",
            [
                ("Corpus size (all postings)", str(corpus_size)),
                ("Matched filters (before per-company cap)", str(matched_before_cap)),
                (
                    "Visible after cap (= rows on 'Jobs' sheet)",
                    str(visible_rows),
                ),
                (
                    "Score range on this sheet",
                    f"{score_min}..{score_max}" if score_min is not None else "(no scored rows)",
                ),
            ],
        ),
        (
            "Active filters",
            [
                ("sort", spec.sort),
                ("per_company_cap", str(spec.per_company_cap)),
                ("tier", _fmt_filter_list(spec.tier)),
                ("ats", _fmt_filter_list(spec.ats)),
                ("remote_type", _fmt_filter_list(spec.remote_type)),
                ("role_family", _fmt_filter_list(spec.role_family)),
                ("state", _fmt_filter_list(spec.state)),
                ("include_snoozed_past_only", _fmt_bool(spec.include_snoozed_past_only)),
                ("include_closed", _fmt_bool(spec.include_closed)),
                ("include_filtered", _fmt_bool(spec.include_filtered)),
                (
                    "target_company_id",
                    str(spec.target_company_id) if spec.target_company_id else "(none)",
                ),
            ],
        ),
        (
            "Scorer weights (sum=100)",
            [(k, str(v)) for k, v in _WEIGHTS.items()],
        ),
        (
            "Hard rules (operator profile id=1)",
            [
                ("salary_floor_usd", f"${hard.salary_floor_usd:,}"),
                (
                    "salary_ceiling_usd",
                    f"${hard.salary_ceiling_usd:,}" if hard.salary_ceiling_usd else "(none)",
                ),
                (
                    "seniority_levels_included",
                    ", ".join(hard.seniority_levels_included) or "(all)",
                ),
                ("geo_whitelist", ", ".join(hard.geo_whitelist) or "(empty)"),
                ("applicant_cap", str(hard.applicant_cap)),
                (
                    "staffing_firm_blocklist",
                    ", ".join(hard.staffing_firm_blocklist) or "(empty)",
                ),
            ],
        ),
        (
            "Reading the fit_score",
            [
                (
                    "Composition",
                    "Weighted sum of five 0-100 sub-scores: role_family, "
                    "seniority, salary, tier, geo (weights above).",
                ),
                (
                    "Role-family gate",
                    "Postings whose role_family is NOT in "
                    f"{sorted(PREFERRED_FAMILIES)} are capped at score=40, "
                    "so every genuine PM role outranks them. See "
                    "services/scoring.py for the gate logic.",
                ),
                (
                    "Per-company cap",
                    f"per_company_cap={spec.per_company_cap}: at most this many "
                    "rows survive per company, ranked inside the bucket by "
                    "score DESC, first_seen DESC, id ASC — regardless of "
                    "the outer sort. So sort=oldest + cap=3 means 'oldest "
                    "of each company's top-3 by score.'",
                ),
                (
                    "Closed / filtered",
                    "Closed = ATS stopped surfacing the role (>=7 days). "
                    "Filtered = posting failed a hard rule. Both hidden by "
                    "default; the include_* flags above show the toggle.",
                ),
            ],
        ),
    ]

    row = 1
    for section_title, items in sections:
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.cell(row=row, column=2, value="").fill = _HEADER_FILL
        row += 1
        for label, value in items:
            kc = ws.cell(row=row, column=1, value=label)
            kc.font = _BOLD
            kc.alignment = _TOP
            vc = ws.cell(row=row, column=2, value=value)
            vc.alignment = _WRAP
            row += 1
        row += 1  # blank spacer


# Column layout for Sheet 2 — (header, width, wrap)
_JOB_COLUMNS: list[tuple[str, int, bool]] = [
    ("rank", 6, False),
    ("company", 28, False),
    ("role", 36, False),
    ("fit_score", 9, False),
    ("score.role_family", 11, False),
    ("score.seniority", 10, False),
    ("score.salary", 9, False),
    ("score.tier", 8, False),
    ("score.geo", 8, False),
    ("role_family", 18, False),
    ("seniority", 14, False),
    ("salary_min", 11, False),
    ("salary_max", 11, False),
    ("salary_currency", 8, False),
    ("location", 28, True),
    ("remote_type", 10, False),
    ("tier", 6, False),
    ("ats_source", 12, False),
    ("apply_url", 50, False),
    ("first_seen_at", 20, False),
    ("jd_summary_markdown", 90, True),
]


def _flatten_locations(locations_normalized: Any) -> str:
    if not isinstance(locations_normalized, list):
        return ""
    out: list[str] = []
    for entry in locations_normalized:
        if not isinstance(entry, dict):
            continue
        for key in ("city", "region", "country", "raw"):
            val = entry.get(key)
            if isinstance(val, str) and val:
                out.append(val)
                break
    return "; ".join(out)


def _build_jobs_sheet(
    ws: Any,
    *,
    rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
    profile: OperatorProfile,
) -> tuple[int | None, int | None]:
    """Sheet 2: one row per posting. Returns (min_score, max_score) seen."""
    ws.title = "Jobs"
    for idx, (header, width, _) in enumerate(_JOB_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    scores: list[int] = []
    for rank, (jp, tc, ps_ats, ps_apply_url) in enumerate(rows, start=1):
        tier = tc.tier if tc is not None else None
        breakdown = score_breakdown(jp, profile, tier=tier)
        company_name = tc.name if tc is not None else jp.canonical_company_name
        first_seen = jp.first_seen_at.isoformat() if jp.first_seen_at else ""
        if jp.fit_score is not None:
            scores.append(int(jp.fit_score))

        values: list[Any] = [
            rank,
            company_name,
            jp.normalized_title,
            jp.fit_score,
            breakdown["role_family"],
            breakdown["seniority"],
            breakdown["salary"],
            breakdown["tier"],
            breakdown["geo"],
            _enum_value(jp.role_family),
            _enum_value(jp.seniority_level),
            jp.salary_min,
            jp.salary_max,
            jp.salary_currency,
            _flatten_locations(jp.locations_normalized) or (jp.location_raw or ""),
            _enum_value(jp.remote_type),
            tier,
            ps_ats or "",
            ps_apply_url or "",
            first_seen,
            jp.jd_summary_markdown or "",
        ]
        for col, (val, (_, _, wrap)) in enumerate(zip(values, _JOB_COLUMNS, strict=True), start=1):
            cell = ws.cell(row=rank + 1, column=col, value=val)
            cell.alignment = _WRAP if wrap else _TOP

    return (min(scores), max(scores)) if scores else (None, None)


def build_workbook_bytes(
    *,
    spec: PostingsViewSpec,
    profile: OperatorProfile,
    rows: list[tuple[JobPosting, TargetCompany | None, str | None, str | None]],
    corpus_size: int,
    matched_before_cap: int,
) -> bytes:
    """Build the two-sheet xlsx and return its bytes."""
    wb = Workbook()
    jobs_ws = wb.active
    # Build Jobs first to get score range, then Context referencing it.
    score_min, score_max = _build_jobs_sheet(jobs_ws, rows=rows, profile=profile)
    context_ws = wb.create_sheet(title="Export Context", index=0)
    _build_context_sheet(
        context_ws,
        spec=spec,
        profile=profile,
        corpus_size=corpus_size,
        matched_before_cap=matched_before_cap,
        visible_rows=len(rows),
        score_min=score_min,
        score_max=score_max,
    )
    # Default-open on the context sheet so the reviewer sees provenance first.
    wb.active = wb.index(context_ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["EXPORT_ROW_CAP", "build_workbook_bytes"]

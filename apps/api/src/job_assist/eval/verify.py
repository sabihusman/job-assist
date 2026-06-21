"""Phase 2 — the human verify surface (Excel) + override-rate scoring.

Offline only; no ``openai`` import (isolation guard stays green). Two pure
entry points used by the runner:

  * ``build_workbook(prelabels)`` — an .xlsx for the operator to verify/correct.
  * ``score(prelabels, jd_rows, email_rows)`` — override rates + verified labels.

Anti-anchoring (operator-approved, strongest form): for the two blanked
dimensions — ``hard_seniority_mismatch`` seniority and rejection-stage
``outcome_type`` — the ``verified_*`` cell is left EMPTY and **no o3/Gemini
anchor appears in the sheet at all**. The o3 label lives only in the pre-label
JSONL, which the scorer reads to compute the unanchored override rate. For all
other rows the visible anchor is simply the pre-filled ``verified_*`` value.
"""

from __future__ import annotations

import hashlib
import json
from collections.abc import Callable
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from job_assist.db.enums import OutcomeType, RoleFamily, SeniorityLevel

NA_NON_PM = "n/a_non_pm"
PM_FAMILIES = frozenset({RoleFamily.product_management.value, RoleFamily.product_owner.value})
REJECTION_STAGES = frozenset(
    {
        OutcomeType.rejection_pre_screen.value,
        OutcomeType.rejection_post_screen.value,
        OutcomeType.rejection_post_interview.value,
    }
)

ROLE_FAMILY_OPTIONS = [m.value for m in RoleFamily]
SENIORITY_OPTIONS = [m.value for m in SeniorityLevel] + [NA_NON_PM]
OUTCOME_OPTIONS = [m.value for m in OutcomeType]

JD_HEADERS = [
    "id",
    "stratum",
    "title",
    "jd_text",
    "verified_role_family",
    "verified_seniority",
    "notes",
]
EMAIL_HEADERS = ["id", "stratum", "subject", "raw_snippet", "verified_outcome_type", "notes"]

_HEADER_FONT = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_UNLOCKED = Protection(locked=False)


def _is_rejection(rec: dict[str, Any]) -> bool:
    return str(rec.get("production_outcome_type")) in REJECTION_STAGES


def _is_seniority_blanked(rec: dict[str, Any]) -> bool:
    return rec.get("stratum") == "hard_seniority_mismatch"


# ── Build ────────────────────────────────────────────────────────────────────


def _add_lists_sheet(wb: Workbook) -> None:
    """Hidden sheet holding dropdown option lists (range-referenced DV avoids the
    255-char inline-list limit)."""
    ws = wb.create_sheet("_lists")
    ws["A1"] = "role_family"
    ws["B1"] = "seniority"
    ws["C1"] = "outcome_type"
    for i, v in enumerate(ROLE_FAMILY_OPTIONS, start=2):
        ws[f"A{i}"] = v
    for i, v in enumerate(SENIORITY_OPTIONS, start=2):
        ws[f"B{i}"] = v
    for i, v in enumerate(OUTCOME_OPTIONS, start=2):
        ws[f"C{i}"] = v
    ws.sheet_state = "hidden"


def _dv(formula1: str) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True, showDropDown=False)
    dv.error = "Pick a value from the dropdown."
    dv.errorTitle = "Invalid value"
    return dv


def _write_header(ws: Any, headers: list[str]) -> None:
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = _HEADER_FONT


def _build_jd_tab(wb: Workbook, jd: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("JDs")
    _write_header(ws, JD_HEADERS)
    rf_dv = _dv(f"=_lists!$A$2:$A${1 + len(ROLE_FAMILY_OPTIONS)}")
    sen_dv = _dv(f"=_lists!$B$2:$B${1 + len(SENIORITY_OPTIONS)}")
    ws.add_data_validation(rf_dv)
    ws.add_data_validation(sen_dv)

    for r, rec in enumerate(jd, start=2):
        label = rec.get("openai_label") or {}
        inp = rec.get("input") or {}
        ws.cell(row=r, column=1, value=rec.get("id"))
        ws.cell(row=r, column=2, value=rec.get("stratum"))
        ws.cell(row=r, column=3, value=inp.get("title"))
        jd_cell = ws.cell(row=r, column=4, value=inp.get("jd_text"))
        jd_cell.alignment = _WRAP
        # verified_role_family — always pre-filled = o3 (visible anchor).
        vrf = ws.cell(row=r, column=5, value=label.get("role_family"))
        # verified_seniority — pre-filled = o3 EXCEPT blanked for the mismatch
        # stratum (cold label, no in-sheet anchor).
        vsen_value = None if _is_seniority_blanked(rec) else label.get("seniority_level")
        vsen = ws.cell(row=r, column=6, value=vsen_value)
        notes = ws.cell(row=r, column=7)
        for c in (vrf, vsen, notes):
            c.protection = _UNLOCKED
        rf_dv.add(vrf)
        sen_dv.add(vsen)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 90
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.freeze_panes = "A2"
    ws.protection.sheet = True


def _build_email_tab(wb: Workbook, emails: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Emails")
    _write_header(ws, EMAIL_HEADERS)
    ot_dv = _dv(f"=_lists!$C$2:$C${1 + len(OUTCOME_OPTIONS)}")
    ws.add_data_validation(ot_dv)

    for r, rec in enumerate(emails, start=2):
        label = rec.get("openai_label") or {}
        inp = rec.get("input") or {}
        ws.cell(row=r, column=1, value=rec.get("id"))
        ws.cell(row=r, column=2, value=rec.get("stratum"))
        ws.cell(row=r, column=3, value=inp.get("subject"))
        snip = ws.cell(row=r, column=4, value=inp.get("raw_snippet"))
        snip.alignment = _WRAP
        # verified_outcome_type — pre-filled = o3 EXCEPT blanked for rejection
        # stages (cold label, no in-sheet anchor).
        vot_value = None if _is_rejection(rec) else label.get("outcome_type")
        vot = ws.cell(row=r, column=5, value=vot_value)
        notes = ws.cell(row=r, column=6)
        for c in (vot, notes):
            c.protection = _UNLOCKED
        ot_dv.add(vot)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 90
    ws.column_dimensions["E"].width = 26
    ws.freeze_panes = "A2"
    ws.protection.sheet = True


def _build_instructions_tab(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions", 0)
    lines = [
        "A4 eval — verify the o3 pre-labels (Phase 2).",
        "",
        "Goal: confirm or CORRECT each label. Read the full source text (jd_text / subject+raw_snippet)",
        "before deciding. Do NOT rubber-stamp — the override rate is the credibility signal.",
        "",
        "JDs tab:",
        "  - verified_role_family: pre-filled with o3's call. Change it where wrong (dropdown).",
        "  - verified_seniority: pre-filled with o3's call EXCEPT the 'hard_seniority_mismatch' rows,",
        "    which are BLANK on purpose — label those from scratch, no anchor shown.",
        "  - NON-PM RULE: if verified_role_family is NOT product_management/product_owner, set",
        f"    verified_seniority to '{NA_NON_PM}'. Seniority is a PM ladder; it does not apply to non-PM",
        "    roles, so those are excluded from the seniority metric (role_family still counts).",
        "",
        "Emails tab:",
        "  - verified_outcome_type: pre-filled with o3's call EXCEPT rejection-stage rows, which are",
        "    BLANK on purpose — label the rejection stage from scratch, no anchor shown.",
        "",
        "Only the verified_* and notes columns are editable; source columns are locked (no password —",
        "you can unprotect if needed). Save and hand back the file; the scorer computes override rates",
        "and emits the verified ground-truth labels.",
    ]
    for i, text in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=text)
    ws.column_dimensions["A"].width = 110


def build_workbook(prelabels: list[dict[str, Any]]) -> Workbook:
    """Build the verify workbook from the pre-label records."""
    jd = [r for r in prelabels if r.get("kind") == "jd"]
    emails = [r for r in prelabels if r.get("kind") == "email"]
    wb = Workbook()
    # Drop the default sheet; we add our own.
    default = wb.active
    wb.remove(default)
    _build_jd_tab(wb, jd)
    _build_email_tab(wb, emails)
    _add_lists_sheet(wb)
    _build_instructions_tab(wb)
    return wb


# ── Score ────────────────────────────────────────────────────────────────────


def _norm(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def score(
    prelabels: list[dict[str, Any]],
    jd_rows: list[dict[str, Any]],
    email_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Compute override rates + verified labels from the edited sheet rows.

    ``jd_rows`` / ``email_rows`` are header->value dicts read from the sheet.
    The o3 label, input snapshot, and input_sha256 come from ``prelabels``
    (keyed by id) — carried forward VERBATIM so Phase 3 re-scores Gemini on
    byte-identical input.
    """
    pre = {str(r.get("id")): r for r in prelabels}
    verified: list[dict[str, Any]] = []

    # JD role_family
    rf_overrides = rf_incomplete = 0
    rf_confusion: dict[str, dict[str, int]] = {}
    # JD seniority (eligible = verified PM-family AND filled AND not n/a)
    sen_overrides = sen_eligible = 0
    sen_excluded_non_pm = sen_excluded_na = sen_incomplete = 0
    sen_confusion: dict[str, dict[str, int]] = {}

    for row in jd_rows:
        rid = str(row.get("id"))
        rec = pre.get(rid)
        if rec is None:
            continue
        o3 = rec.get("openai_label") or {}
        o3_rf = _norm(o3.get("role_family"))
        o3_sen = _norm(o3.get("seniority_level"))
        v_rf = _norm(row.get("verified_role_family"))
        v_sen = _norm(row.get("verified_seniority"))

        # role_family
        if v_rf is None:
            rf_incomplete += 1
            rf_overridden: bool | None = None
        else:
            rf_overridden = v_rf != o3_rf
            rf_overrides += int(rf_overridden)
            rf_confusion.setdefault(str(o3_rf), {}).setdefault(v_rf, 0)
            rf_confusion[str(o3_rf)][v_rf] += 1

        # seniority eligibility / override
        eligible = v_rf in PM_FAMILIES and v_sen is not None and v_sen != NA_NON_PM
        sen_overridden: bool | None
        if v_sen == NA_NON_PM:
            sen_excluded_na += 1
            sen_overridden = None
        elif v_rf is not None and v_rf not in PM_FAMILIES:
            sen_excluded_non_pm += 1
            sen_overridden = None
        elif v_sen is None:
            sen_incomplete += 1
            sen_overridden = None
        else:
            sen_eligible += 1
            sen_overridden = v_sen != o3_sen
            sen_overrides += int(sen_overridden)
            sen_confusion.setdefault(str(o3_sen), {}).setdefault(v_sen, 0)
            sen_confusion[str(o3_sen)][v_sen] += 1

        verified.append(
            {
                "kind": "jd",
                "id": rid,
                "stratum": rec.get("stratum"),
                "input": rec.get("input"),
                "input_sha256": rec.get("input_sha256"),
                "o3_label": o3,
                "verified_label": {
                    "role_family": v_rf,
                    "seniority_level": None if (v_sen == NA_NON_PM) else v_sen,
                },
                "seniority_eval_eligible": eligible,
                "was_overridden_role_family": rf_overridden,
                "was_overridden_seniority": sen_overridden,
            }
        )

    # Email outcome_type
    ot_overrides = ot_eval = ot_incomplete = 0
    ot_confusion: dict[str, dict[str, int]] = {}
    for row in email_rows:
        rid = str(row.get("id"))
        rec = pre.get(rid)
        if rec is None:
            continue
        o3 = rec.get("openai_label") or {}
        o3_ot = _norm(o3.get("outcome_type"))
        v_ot = _norm(row.get("verified_outcome_type"))
        if v_ot is None:
            ot_incomplete += 1
            ot_overridden: bool | None = None
        else:
            ot_eval += 1
            ot_overridden = v_ot != o3_ot
            ot_overrides += int(ot_overridden)
            ot_confusion.setdefault(str(o3_ot), {}).setdefault(v_ot, 0)
            ot_confusion[str(o3_ot)][v_ot] += 1
        verified.append(
            {
                "kind": "email",
                "id": rid,
                "stratum": rec.get("stratum"),
                "production_outcome_type": rec.get("production_outcome_type"),
                "input": rec.get("input"),
                "input_sha256": rec.get("input_sha256"),
                "o3_label": o3,
                "verified_label": {"outcome_type": v_ot},
                "was_overridden": ot_overridden,
            }
        )

    def _rate(n: int, d: int) -> float | None:
        return round(n / d, 4) if d else None

    sample_model = next((r.get("model_id") for r in prelabels if r.get("model_id")), None)
    summary = {
        "model_id": sample_model,
        "jd": {
            "n_rows": sum(1 for r in jd_rows if pre.get(str(r.get("id")))),
            "role_family": {
                "overrides": rf_overrides,
                "scored": sum(sum(v.values()) for v in rf_confusion.values()),
                "override_rate": _rate(
                    rf_overrides, sum(sum(v.values()) for v in rf_confusion.values())
                ),
                "incomplete_blank": rf_incomplete,
                "confusion_o3_to_verified": rf_confusion,
            },
            "seniority": {
                "n_eligible": sen_eligible,
                "overrides": sen_overrides,
                "override_rate": _rate(sen_overrides, sen_eligible),
                "excluded_non_pm": sen_excluded_non_pm,
                "excluded_na_non_pm": sen_excluded_na,
                "incomplete_blank": sen_incomplete,
                "confusion_o3_to_verified": sen_confusion,
            },
        },
        "email": {
            "n_rows": sum(1 for r in email_rows if pre.get(str(r.get("id")))),
            "outcome_type": {
                "scored": ot_eval,
                "overrides": ot_overrides,
                "override_rate": _rate(ot_overrides, ot_eval),
                "incomplete_blank": ot_incomplete,
                "confusion_o3_to_verified": ot_confusion,
            },
        },
    }
    return verified, summary


# ── Sheet IO (thin; the math above is pure for testing) ──────────────────────


def _input_sha256(payload: dict[str, Any]) -> str:
    """Match run._sha256 exactly so reconstructed hashes equal the originals."""
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def finalize(
    build_jd_rows: list[dict[str, Any]],
    build_em_rows: list[dict[str, Any]],
    corr_jd_rows: list[dict[str, Any]],
    corr_em_rows: list[dict[str, Any]],
    *,
    relabel_jd: Callable[[str, str], str | None],
    relabel_em: Callable[[str, str], str | None],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    """Recover the complete o3 baseline and score the corrected sheet against it.

    o3 for the 109 non-anchor rows comes from the ORIGINAL build sheet's
    pre-fills (the labels the operator anchored on). o3 for the 47 anti-anchor
    rows (blank in the build sheet, lost with the JSONL) is recovered by a FRESH
    o3 relabel of just those inputs — valid because those rows were labeled cold,
    so a fresh independent o3 is the intended unanchored comparison.

    ``relabel_jd(title, jd_text) -> seniority_level`` and
    ``relabel_em(subject, raw_snippet) -> outcome_type`` are injected (run.py
    wires them to the o3 labeler; tests pass stubs). They are called ONLY for
    rows missing an o3 label.
    """
    b_jd = {str(r["id"]): r for r in build_jd_rows}
    b_em = {str(r["id"]): r for r in build_em_rows}

    # Guard: a PRISTINE build sheet always leaves the anti-anchor strata blank
    # (hard_seniority_mismatch seniority + rejection-stage outcome). If the
    # passed build sheet has ZERO such blanks it's a filled/corrected copy, not
    # the original o3 build — comparing against it silently yields 0% override.
    # Fail loud instead.
    build_blanks = sum(
        1 for r in build_jd_rows if _norm(r.get("verified_seniority")) is None
    ) + sum(1 for r in build_em_rows if _norm(r.get("verified_outcome_type")) is None)
    if build_blanks == 0:
        raise ValueError(
            "--build-xlsx has no anti-anchor blanks (0 blank verified_seniority / "
            "verified_outcome_type cells). That looks like a filled or corrected "
            "copy, not the pristine o3 build sheet — pass the ORIGINAL build sheet "
            "(the one verify-build produced, with the mismatch seniority + "
            "rejection outcome cells left blank)."
        )

    prelabels: list[dict[str, Any]] = []
    relabeled = 0

    for r in corr_jd_rows:
        rid = str(r["id"])
        b = b_jd.get(rid, {})
        o3_rf = _norm(b.get("verified_role_family"))  # build always pre-filled rf
        o3_sen = _norm(b.get("verified_seniority"))  # None for the 20 mismatch rows
        title = r.get("title") or ""
        jd_text = r.get("jd_text") or ""
        source = "build_prefill"
        if o3_sen is None:
            o3_sen = _norm(relabel_jd(title, jd_text))
            relabeled += 1
            source = "fresh_relabel"
        inp = {"title": r.get("title"), "jd_text": r.get("jd_text")}
        prelabels.append(
            {
                "kind": "jd",
                "id": rid,
                "stratum": r.get("stratum"),
                "input": inp,
                "input_sha256": _input_sha256(inp),
                "openai_label": {"role_family": o3_rf, "seniority_level": o3_sen},
                "o3_source": source,
            }
        )

    for r in corr_em_rows:
        rid = str(r["id"])
        b = b_em.get(rid, {})
        o3_ot = _norm(b.get("verified_outcome_type"))  # None for the 27 rejection rows
        subject = r.get("subject") or ""
        snippet = r.get("raw_snippet") or ""
        source = "build_prefill"
        if o3_ot is None:
            o3_ot = _norm(relabel_em(subject, snippet))
            relabeled += 1
            source = "fresh_relabel"
        inp = {"subject": r.get("subject"), "raw_snippet": r.get("raw_snippet")}
        prelabels.append(
            {
                "kind": "email",
                "id": rid,
                "stratum": r.get("stratum"),
                "production_outcome_type": r.get("stratum"),
                "input": inp,
                "input_sha256": _input_sha256(inp),
                "openai_label": {"outcome_type": o3_ot},
                "o3_source": source,
            }
        )

    verified, summary = score(prelabels, corr_jd_rows, corr_em_rows)
    summary["relabeled_anchor_rows"] = relabeled
    summary["recovery"] = "o3: 109 build-sheet prefills + fresh relabel of anti-anchor rows"
    return prelabels, verified, summary


def read_verify_rows(wb: Workbook) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read the edited workbook back into header->value row dicts."""

    def _rows(sheet_name: str, headers: list[str]) -> list[dict[str, Any]]:
        ws = wb[sheet_name]
        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            out.append({h: row[i] if i < len(row) else None for i, h in enumerate(headers)})
        return out

    return _rows("JDs", JD_HEADERS), _rows("Emails", EMAIL_HEADERS)

"""Offline eval runner (CLI) — NEVER a route/cron/sweep.

Modes:
  * ``count``         — read-only: pull AGGREGATE prod counts for sample sizing
                        (no per-row text). Safe for CI (``eval-count`` workflow)
                        or local.
  * ``generate``      — o3 pre-labeler over the stratified sample → pre-label
                        JSONL. LOCAL ONLY (real JD text + email snippets).
  * ``verify-build``  — build the Excel verify sheet. LOCAL ONLY.
  * ``verify-score``  — override rates + verified labels. LOCAL ONLY.

PUBLIC repo: only ``count`` runs in CI (aggregates only). The data-bearing
modes have no CI workflow and run on the operator's machine with OPENAI_API_KEY
/ API_URL / API_AUTH_TOKEN in the env. All outputs under ``eval/datasets/`` and
``verify_inbox/`` are gitignored. See ``eval/README.md``.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from job_assist.eval.api_client import (
    fetch_open_postings,
    fetch_outcome_type_breakdown,
    fetch_outcomes,
    fetch_posting_detail,
)
from job_assist.eval.sample import compute_counts, select_email_sample, select_jd_sample

DATASETS_DIR = Path(__file__).parent / "datasets"


def _sha256(payload: dict[str, Any]) -> str:
    """Stable hash of the exact model input — Phase 3 asserts identical input."""
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()


def _dedup_by_id(*lists: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    out: list[dict[str, Any]] = []
    for lst in lists:
        for row in lst:
            rid = str(row.get("id"))
            if rid not in seen:
                seen.add(rid)
                out.append(row)
    return out


def _run_count(stamp: str) -> int:
    """Read-only count step. Prints JSON and writes a timestamped artifact."""
    postings = fetch_open_postings()
    counts = compute_counts(postings)
    counts["by_outcome_type"] = fetch_outcome_type_breakdown()
    counts["generated_at"] = stamp

    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    out = DATASETS_DIR / f"pool_counts.{stamp}.json"
    out.write_text(json.dumps(counts, indent=2, sort_keys=True), encoding="utf-8")

    # Print to stdout so the workflow log carries the numbers.
    print(json.dumps(counts, indent=2, sort_keys=True))
    print(f"\n[count] wrote {out}", file=sys.stderr)
    return 0


def _run_generate(stamp: str) -> int:
    """Pre-label the confirmed stratified sample with the o-series model.

    Identical-input lock: JDs use ``description_markdown``; emails use
    ``subject`` + ``raw_snippet`` — the SAME text Phase-3 Gemini must re-score,
    captured verbatim with a sha256 so the comparison measures model difference,
    not input difference. Per-item failures are collected, not fatal.
    """
    from job_assist.eval.openai_labeler import label_email, label_jd, new_client

    client = new_client()
    labeled_at = datetime.now(UTC).isoformat()
    records: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    jd_dist: Counter[str] = Counter()
    sen_dist: Counter[str] = Counter()
    email_dist: Counter[str] = Counter()
    # Q3 sanity: for the hard-seniority-mismatch stratum, capture title +
    # Gemini's (production) seniority + o3's seniority side by side.
    mismatch_detail: list[dict[str, Any]] = []

    # ── JDs ──────────────────────────────────────────────────────────────────
    jd_sample = select_jd_sample(fetch_open_postings())
    for item in jd_sample:
        pid = str(item.get("id"))
        try:
            detail = fetch_posting_detail(pid)
            model_input = {"title": detail["title"], "jd_text": detail["jd_text"]}
            res = label_jd(client, title=detail["title"], jd_text=detail["jd_text"])
            records.append(
                {
                    "kind": "jd",
                    "id": pid,
                    "stratum": item.get("_stratum"),
                    "input": model_input,
                    "input_sha256": _sha256(model_input),
                    "openai_label": res.label,
                    "model_id": res.served_model,
                    "prompt_version": res.prompt_version,
                    "temperature_mode": res.temperature_mode,
                    "generated_at": stamp,
                    "labeled_at": labeled_at,
                }
            )
            jd_dist[str(res.label.get("role_family"))] += 1
            sen_dist[str(res.label.get("seniority_level"))] += 1
            if item.get("_stratum") == "hard_seniority_mismatch":
                role = item.get("role") or {}
                mismatch_detail.append(
                    {
                        "title": role.get("title"),
                        "gemini_seniority": role.get("seniority"),
                        "o3_seniority": res.label.get("seniority_level"),
                        "o3_role_family": res.label.get("role_family"),
                    }
                )
        except Exception as exc:  # collect, never abort the batch
            errors.append({"kind": "jd", "id": pid, "error": str(exc)[:300]})

    # ── Emails (identical subject+raw_snippet input) ─────────────────────────
    lifecycle = fetch_outcomes(job_related=True)
    negatives = fetch_outcomes(job_related=False, max_rows=800)
    email_sample = select_email_sample(_dedup_by_id(lifecycle, negatives))
    for o in email_sample:
        oid = str(o.get("id"))
        try:
            subject = o.get("subject") or ""
            snippet = o.get("raw_snippet") or ""
            model_input = {"subject": subject, "raw_snippet": snippet}
            res = label_email(client, subject=subject, body=snippet)
            records.append(
                {
                    "kind": "email",
                    "id": oid,
                    "stratum": o.get("_stratum"),
                    "production_outcome_type": o.get("stage"),
                    "input": model_input,
                    "input_sha256": _sha256(model_input),
                    "openai_label": res.label,
                    "model_id": res.served_model,
                    "prompt_version": res.prompt_version,
                    "temperature_mode": res.temperature_mode,
                    "generated_at": stamp,
                    "labeled_at": labeled_at,
                }
            )
            email_dist[str(res.label.get("outcome_type"))] += 1
        except Exception as exc:
            errors.append({"kind": "email", "id": oid, "error": str(exc)[:300]})

    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    out = DATASETS_DIR / f"prelabels.{stamp}.jsonl"
    with out.open("w", encoding="utf-8") as fh:
        for rec in records:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")

    summary = {
        "jd_labeled": jd_dist.total(),
        "email_labeled": email_dist.total(),
        "total_labeled": len(records),
        "errors": len(errors),
        "o3_jd_role_family_distribution": dict(jd_dist),
        "o3_jd_seniority_distribution": dict(sen_dist),
        "o3_email_outcome_distribution": dict(email_dist),
        "hard_seniority_mismatch_detail": sorted(
            mismatch_detail, key=lambda d: str(d.get("title"))
        ),
        "error_detail": errors[:20],
    }
    print(json.dumps(summary, indent=2, sort_keys=True))
    print(f"\n[generate] wrote {len(records)} records → {out}", file=sys.stderr)
    return 0


def _read_jsonl(path: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            out.append(json.loads(line))
    return out


def _run_verify_build(stamp: str, jsonl: str) -> int:
    """Build the Excel verify sheet from a pre-label JSONL."""
    from job_assist.eval.verify import build_workbook

    prelabels = _read_jsonl(jsonl)
    wb = build_workbook(prelabels)
    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    out = DATASETS_DIR / f"verify_sheet.{stamp}.xlsx"
    wb.save(out)
    jd = sum(1 for r in prelabels if r.get("kind") == "jd")
    em = sum(1 for r in prelabels if r.get("kind") == "email")
    print(json.dumps({"jd_rows": jd, "email_rows": em, "sheet": str(out.name)}, indent=2))
    print(f"\n[verify-build] wrote {out}", file=sys.stderr)
    return 0


def _run_verify_score(stamp: str, jsonl: str, xlsx: str) -> int:
    """Score the edited verify sheet → verified labels JSONL + override summary."""
    from openpyxl import load_workbook

    from job_assist.eval.verify import read_verify_rows, score

    prelabels = _read_jsonl(jsonl)
    wb = load_workbook(xlsx, data_only=True)
    jd_rows, email_rows = read_verify_rows(wb)
    verified, summary = score(prelabels, jd_rows, email_rows)

    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    labels_out = DATASETS_DIR / f"verified_labels.{stamp}.jsonl"
    with labels_out.open("w", encoding="utf-8") as fh:
        for rec in verified:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
    summary_out = DATASETS_DIR / f"override_summary.{stamp}.json"
    summary_out.write_text(json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8")

    print(json.dumps(summary, indent=2, sort_keys=True))
    print(
        f"\n[verify-score] wrote {len(verified)} labels → {labels_out.name}; "
        f"summary → {summary_out.name}",
        file=sys.stderr,
    )
    return 0


def _run_verify_finalize(stamp: str, build_xlsx: str, corrected_xlsx: str) -> int:
    """Recovery path: o3 for the 109 non-anchor rows from the original build
    sheet's prefills + a FRESH o3 relabel of the 47 anti-anchor rows, then score
    the corrected sheet. LOCAL ONLY — needs OPENAI_API_KEY in the env for the
    47-row relabel. Use when the original pre-label JSONL is lost.
    """
    from openpyxl import load_workbook

    from job_assist.eval.openai_labeler import label_email, label_jd, new_client
    from job_assist.eval.verify import finalize, read_verify_rows

    bjd, bem = read_verify_rows(load_workbook(build_xlsx, data_only=True))
    cjd, cem = read_verify_rows(load_workbook(corrected_xlsx, data_only=True))

    # Lazy client: only constructed (and only needs the key) if there are
    # anti-anchor rows to relabel.
    holder: dict[str, Any] = {}

    def _client() -> Any:
        if "c" not in holder:
            holder["c"] = new_client()
        return holder["c"]

    def relabel_jd(title: str, jd_text: str) -> str | None:
        return label_jd(_client(), title=title, jd_text=jd_text).label.get("seniority_level")

    def relabel_em(subject: str, snippet: str) -> str | None:
        return label_email(_client(), subject=subject, body=snippet).label.get("outcome_type")

    _prelabels, verified, summary = finalize(
        bjd, bem, cjd, cem, relabel_jd=relabel_jd, relabel_em=relabel_em
    )

    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    labels_out = DATASETS_DIR / f"verified_labels.{stamp}.jsonl"
    with labels_out.open("w", encoding="utf-8") as fh:
        for rec in verified:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
    (DATASETS_DIR / f"override_summary.{stamp}.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    print(
        f"\n[verify-finalize] relabeled {summary['relabeled_anchor_rows']} anti-anchor rows; "
        f"wrote {len(verified)} verified labels → {labels_out.name}",
        file=sys.stderr,
    )
    return 0


def _run_gemini_score(stamp: str, labels: str, profile_context: str | None) -> int:
    """Score the UNCHANGED production Gemini classifier vs the verified labels.

    LOCAL ONLY — reads the gitignored verified_labels JSONL, runs classify_posting
    (JDs) and GmailOutcomeClassifier.classify (emails) on the SAME input bytes
    (input_sha256 lock), and writes the three-way accuracy summary. Needs
    GEMINI_API_KEY in the env. Observes the classifier; changes nothing.
    """
    import asyncio
    from datetime import UTC, datetime

    from tenacity import retry, retry_if_exception, stop_after_attempt, wait_exponential

    from job_assist.config import settings
    from job_assist.eval.gemini_score import _is_transient_error, aggregate, collect
    from job_assist.gmail.classifier import EmailClassifier
    from job_assist.gmail.models import RawEmail
    from job_assist.services.classifier import classify_posting

    rows = _read_jsonl(labels)
    gmail_clf = EmailClassifier(api_key=settings.gemini_api_key)
    fixed_ts = datetime(2026, 1, 1, tzinfo=UTC)

    # Bounded exponential backoff on transient Gemini errors (503 "high demand"
    # / overload, 429). Wraps ONLY the eval's injected callables — the production
    # classify_posting / EmailClassifier are untouched. Without this a single
    # 503 aborts the ~90-row run; here it retries (4->60s, up to 6 attempts)
    # before collect()'s per-row guard would skip the row.
    _transient_retry = retry(
        retry=retry_if_exception(_is_transient_error),
        wait=wait_exponential(multiplier=2, min=4, max=60),
        stop=stop_after_attempt(6),
        reraise=True,
    )

    @_transient_retry
    async def classify_jd(title: str, jd_text: str) -> tuple[Any, Any]:
        return await classify_posting(jd_text=jd_text, title=title, profile_context=profile_context)

    @_transient_retry
    async def classify_email(subject: str, snippet: str) -> Any:
        # from_address/from_domain are required by the model but unused by the
        # classifier prompt (build_prompt reads subject + body_text only).
        email = RawEmail(
            message_id="eval",
            from_address="eval@local",
            from_domain="local",
            subject=subject,
            received_at=fixed_ts,
            body_text=snippet,
        )
        result = await gmail_clf.classify(email)
        return result.outcome_type

    scored, skipped = asyncio.run(
        collect(rows, classify_jd=classify_jd, classify_email=classify_email)
    )
    summary = aggregate(scored, skipped, profile_context_used=profile_context is not None)

    DATASETS_DIR.mkdir(parents=True, exist_ok=True)
    scores_out = DATASETS_DIR / f"gemini_scores.{stamp}.jsonl"
    with scores_out.open("w", encoding="utf-8") as fh:
        for rec in scored:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
    (DATASETS_DIR / f"gemini_accuracy_summary.{stamp}.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    print(
        f"\n[gemini-score] scored {len(scored)} rows ({len(skipped)} skipped) → "
        f"{scores_out.name} + gemini_accuracy_summary.{stamp}.json",
        file=sys.stderr,
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="eval.run")
    parser.add_argument(
        "mode",
        choices=[
            "count",
            "generate",
            "verify-build",
            "verify-score",
            "verify-finalize",
            "gemini-score",
        ],
    )
    parser.add_argument(
        "--stamp",
        required=True,
        help="Timestamp tag for artifacts (passed in for reproducibility).",
    )
    parser.add_argument("--jsonl", help="Pre-label JSONL path (verify-build / verify-score).")
    parser.add_argument("--xlsx", help="Edited/corrected verify sheet path.")
    parser.add_argument(
        "--build-xlsx",
        help="Original build sheet (o3 prefills) — verify-finalize recovery path.",
    )
    parser.add_argument(
        "--labels",
        help="verified_labels JSONL path (gemini-score).",
    )
    parser.add_argument(
        "--profile-context",
        help=(
            "gemini-score: optional operator disambiguation context to match prod "
            "exactly. Omit for the no-profile classifier path (recorded in the summary)."
        ),
    )
    args = parser.parse_args(argv)
    if args.mode == "count":
        return _run_count(args.stamp)
    if args.mode == "generate":
        return _run_generate(args.stamp)
    if args.mode == "verify-build":
        if not args.jsonl:
            parser.error("verify-build requires --jsonl")
        return _run_verify_build(args.stamp, args.jsonl)
    if args.mode == "verify-finalize":
        if not args.build_xlsx or not args.xlsx:
            parser.error("verify-finalize requires --build-xlsx and --xlsx")
        return _run_verify_finalize(args.stamp, args.build_xlsx, args.xlsx)
    if args.mode == "gemini-score":
        labels = args.labels or str(DATASETS_DIR / "verified_labels.final.jsonl")
        return _run_gemini_score(args.stamp, labels, args.profile_context)
    # verify-score
    if not args.jsonl or not args.xlsx:
        parser.error("verify-score requires --jsonl and --xlsx")
    return _run_verify_score(args.stamp, args.jsonl, args.xlsx)


if __name__ == "__main__":
    raise SystemExit(main())

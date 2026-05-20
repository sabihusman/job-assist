"""Convert the Tippie alumni directory xlsx → seed JSON (PR #39).

Operator-local one-shot. Reads ``Connecting_With_Classmates_Directory.xlsx``
(25 columns documented in the PR #39 spec) and emits the JSON array
that ``POST /admin/seed/contacts`` accepts.

Usage::

    python apps/api/scripts/tippie_to_contacts_json.py \\
        ~/path/to/Connecting_With_Classmates_Directory.xlsx \\
        > /tmp/contacts.json

The script never runs in the Railway container — ``openpyxl`` is a
dev-only dependency. The output is the JSON body of the seed endpoint;
the operator pipes it into curl. The source xlsx and the resulting JSON
are both gitignored (see repo root ``.gitignore``); the script ALWAYS
writes JSON to stdout and logs to stderr so nothing PII-containing
hits a tracked file unless the operator redirects it themselves.

Skip rules:
    * both first_name and last_name empty   → skip
    * neither email_primary nor linkedin_url → skip (would violate
      the DB CHECK constraint ``ck_contact_has_channel``)

Logging (stderr): counts only — no names, emails, or LinkedIn URLs.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

# Column indices (1-based per spec; we convert to 0-based for openpyxl
# tuple indexing). Kept as constants here so the test fixture and the
# script stay in lockstep.
COL_FORM_ID = 1
COL_EMAIL = 4
COL_NAME = 5
COL_PREFERRED_FIRST = 6
COL_LAST_NAME = 7
COL_UIOWA_EMAIL = 8
COL_PERSONAL_EMAIL = 9
COL_LINKEDIN = 10
COL_CITY = 11
COL_STATE = 12
COL_COUNTRY = 13
COL_METRO = 14
COL_TIPPIE_PROGRAMS = 15
COL_SEMESTER_START = 16
COL_GRADUATION_SEMESTER = 17
COL_EMPLOYER = 18
COL_POSITION = 19
COL_JOB_FUNCTIONS = 20
COL_INDUSTRIES = 21
COL_PREV_UNDERGRAD = 22
COL_PREV_GRADUATE = 23
COL_HOBBIES = 24
COL_TOPICS = 25


def _cell(row: tuple[Any, ...], col_1based: int) -> Any:
    """Safe 1-based cell access. Returns ``None`` for out-of-range cols."""
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _str(value: Any) -> str | None:
    """Coerce + strip; non-strings → str(); empty / 'None' → ``None``."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _split_semis(value: Any) -> list[str] | None:
    """Split semicolon-delimited cell → list of non-empty trimmed items.

    Returns ``None`` if the cell is empty so the Pydantic schema's
    "list or null" contract is preserved.
    """
    text = _str(value)
    if text is None:
        return None
    parts = [p.strip() for p in text.split(";")]
    parts = [p for p in parts if p]
    return parts or None


def _derive_first_last(row: tuple[Any, ...]) -> tuple[str | None, str | None]:
    """Prefer (preferred_first or first-token-of-name, col 7 last_name).

    Col 7 ("Last Name") is the authoritative last_name when present.
    For first_name we prefer col 6 ("Preferred First Name"); if blank,
    we split col 5 ("Name") on whitespace and take the first token.
    Returns ``(first, last)``; either may be ``None`` if the source row
    is too sparse — the caller skips those rows.
    """
    preferred = _str(_cell(row, COL_PREFERRED_FIRST))
    last = _str(_cell(row, COL_LAST_NAME))
    full = _str(_cell(row, COL_NAME))

    first: str | None = preferred
    if first is None and full is not None:
        tokens = full.split()
        first = tokens[0] if tokens else None

    if last is None and full is not None:
        tokens = full.split()
        if len(tokens) >= 2:
            last = tokens[-1]

    return first, last


def _pick_primary_email(row: tuple[Any, ...]) -> str | None:
    """Col 4 (form Email) first, then @uiowa, then personal."""
    for col in (COL_EMAIL, COL_UIOWA_EMAIL, COL_PERSONAL_EMAIL):
        value = _str(_cell(row, col))
        if value is not None:
            return value
    return None


def _build_source_metadata(row: tuple[Any, ...]) -> dict[str, Any]:
    """Collect Tippie-specific extras into the source_metadata JSONB.

    Keys with empty values are dropped — the JSONB column stays compact
    and the operator can grep it later without sifting through ``null``s.
    """
    metadata: dict[str, Any] = {}
    programs = _split_semis(_cell(row, COL_TIPPIE_PROGRAMS))
    if programs is not None:
        metadata["tippie_programs"] = programs
    for key, col in (
        ("semester_start", COL_SEMESTER_START),
        ("graduation_semester", COL_GRADUATION_SEMESTER),
        ("previous_undergrad", COL_PREV_UNDERGRAD),
        ("previous_graduate", COL_PREV_GRADUATE),
        ("hobbies", COL_HOBBIES),
    ):
        value = _str(_cell(row, col))
        if value is not None:
            metadata[key] = value
    form_id = _cell(row, COL_FORM_ID)
    if form_id is not None and _str(form_id) is not None:
        # form_id stays as the original type (often int) so the operator
        # can correlate back to the source spreadsheet without coercion.
        metadata["form_id"] = form_id
    return metadata


def row_to_contact(row: tuple[Any, ...]) -> dict[str, Any] | None:
    """Convert one xlsx row → ContactSeedRow-shaped dict (or ``None`` to skip).

    Pure function — no I/O, no module-level state — so the tests can
    feed it synthetic tuples without spinning up openpyxl. Returns None
    when the row lacks a name or any contact channel.
    """
    first_name, last_name = _derive_first_last(row)
    if first_name is None and last_name is None:
        return None

    email_primary = _pick_primary_email(row)
    linkedin_url = _str(_cell(row, COL_LINKEDIN))
    if email_primary is None and linkedin_url is None:
        return None

    # Fill missing first/last with "" — the Pydantic validator on the
    # API side rejects empty names, so a row with one half missing
    # still gets reported as skipped_invalid. We could drop the row
    # here instead, but letting it through preserves the existing
    # skipped_invalid signal the operator may use to spot-fix the
    # source spreadsheet.
    contact: dict[str, Any] = {
        "first_name": first_name or "",
        "last_name": last_name or "",
        "preferred_first_name": _str(_cell(row, COL_PREFERRED_FIRST)),
        "email_primary": email_primary,
        "email_secondary": _str(_cell(row, COL_PERSONAL_EMAIL))
        if email_primary != _str(_cell(row, COL_PERSONAL_EMAIL))
        else None,
        "linkedin_url": linkedin_url,
        "current_employer": _str(_cell(row, COL_EMPLOYER)),
        "current_position": _str(_cell(row, COL_POSITION)),
        "location_city": _str(_cell(row, COL_CITY)),
        "location_state": _str(_cell(row, COL_STATE)),
        "location_country": _str(_cell(row, COL_COUNTRY)),
        "location_metro": _str(_cell(row, COL_METRO)),
        "source_type": "tippie_alumni",
        "source_metadata": _build_source_metadata(row),
        "job_functions_of_interest": _split_semis(_cell(row, COL_JOB_FUNCTIONS)),
        "industries_of_interest": _split_semis(_cell(row, COL_INDUSTRIES)),
    }

    topics = _split_semis(_cell(row, COL_TOPICS))
    if topics:
        contact["contact_opt_in"] = True
        contact["contact_opt_in_topics"] = topics
    else:
        contact["contact_opt_in"] = False
        contact["contact_opt_in_topics"] = None

    return contact


def convert_workbook(path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Load the xlsx and emit (rows, skip_counts).

    Header row (row 1) is skipped. ``openpyxl`` is imported lazily so
    the unit tests can ``row_to_contact`` directly without the dep.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:  # pragma: no cover — load_workbook returns at least one
        raise RuntimeError("workbook has no active sheet")

    rows: list[dict[str, Any]] = []
    skipped_no_name = 0
    skipped_no_channel = 0

    for i, raw_row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if raw_row is None:
            continue
        first_name, last_name = _derive_first_last(raw_row)
        if first_name is None and last_name is None:
            skipped_no_name += 1
            continue
        contact = row_to_contact(raw_row)
        if contact is None:
            skipped_no_channel += 1
            continue
        rows.append(contact)

    wb.close()
    return rows, {
        "skipped_no_name": skipped_no_name,
        "skipped_no_channel": skipped_no_channel,
    }


def _main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert Tippie alumni xlsx → /admin/seed/contacts JSON",
    )
    parser.add_argument(
        "xlsx",
        type=Path,
        help="Path to Connecting_With_Classmates_Directory.xlsx",
    )
    args = parser.parse_args(argv)

    if not args.xlsx.exists():
        print(f"error: file not found: {args.xlsx}", file=sys.stderr)
        return 2

    rows, counts = convert_workbook(args.xlsx)

    # stderr: counts only — never PII.
    print(
        f"Converted {len(rows)} contacts "
        f"(skipped {counts['skipped_no_name']} for missing name, "
        f"{counts['skipped_no_channel']} for missing contact channel).",
        file=sys.stderr,
    )

    json.dump(rows, sys.stdout, indent=2, ensure_ascii=False, default=str)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
